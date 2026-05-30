"""Generacion de reportes Excel."""
import io

from ..units import convertir_a_cm, convertir_desde_cm, obtener_simbolo_area, obtener_simbolo_unidad

def generar_excel(optimizacion, info_desperdicio, piezas_con_nombre, numero_lista=None):
    """
    Genera un archivo Excel con la información detallada de la optimización.
    
    Args:
        optimizacion: Objeto Optimizacion
        info_desperdicio: Diccionario con información de desperdicio
        piezas_con_nombre: Lista de diccionarios con información de piezas
        numero_lista: Número de lista de la optimización (opcional)
    
    Returns:
        BytesIO: Buffer con el archivo Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # === HOJA 1: RESUMEN GENERAL ===
    ws1 = wb.active
    ws1.title = "Resumen"
    
    # Título
    ws1['A1'] = f"Optimización #{numero_lista if numero_lista else optimizacion.id}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:B1')
    
    # Información general
    row = 3
    # Calcular porcentaje de desperdicio
    area_total_disponible = info_desperdicio.get('area_usada_total', 0) + info_desperdicio.get('desperdicio_total', 0)
    porcentaje_desperdicio = (info_desperdicio.get('desperdicio_total', 0) / area_total_disponible * 100) if area_total_disponible > 0 else 0
    
    datos_generales = [
        ['Fecha:', optimizacion.fecha.strftime('%d/%m/%Y %H:%M')],
        ['Hora:', optimizacion.fecha.strftime('%H:%M') + ' (Chile)'],
        ['Tablero:', f"{round(convertir_desde_cm(optimizacion.ancho_tablero, optimizacion.unidad_medida), 2)} × {round(convertir_desde_cm(optimizacion.alto_tablero, optimizacion.unidad_medida), 2)} {obtener_simbolo_unidad(optimizacion.unidad_medida)}"],
        ['Aprovechamiento:', f"{optimizacion.aprovechamiento_total:.2f}%"],
        ['Área Utilizada:', f"{info_desperdicio.get('area_usada_total', 0)} {obtener_simbolo_area(optimizacion.unidad_medida)}"],
        ['Desperdicio Total:', f"{info_desperdicio.get('desperdicio_total', 0)} {obtener_simbolo_area(optimizacion.unidad_medida)}"],
        ['% Desperdicio:', f"{porcentaje_desperdicio:.2f}%"],
        ['Total Tableros:', len(info_desperdicio.get('info_tableros', []))],
        ['Rotación Automática:', 'Sí' if getattr(optimizacion, 'permitir_rotacion', True) else 'No'],
        ['Margen de Corte:', f"{round(getattr(optimizacion, 'margen_corte', 0.3) * 10, 1)} mm"],
    ]
    
    for label, valor in datos_generales:
        ws1[f'A{row}'] = label
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = valor
        row += 1
    
    # Información de costos (si existe)
    costo_total = optimizacion.get_costo_total()
    if costo_total or optimizacion.precio_tablero or optimizacion.mano_obra:
        row += 1
        ws1[f'A{row}'] = "--- Información de Costos ---"
        ws1[f'A{row}'].font = Font(bold=True, size=12)
        ws1[f'A{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        row += 1
        
        if optimizacion.material:
            ws1[f'A{row}'] = 'Material:'
            ws1[f'A{row}'].font = Font(bold=True)
            ws1[f'B{row}'] = optimizacion.material.nombre
            row += 1
        
        if optimizacion.precio_tablero:
            ws1[f'A{row}'] = 'Precio por Tablero:'
            ws1[f'A{row}'].font = Font(bold=True)
            ws1[f'B{row}'] = f"${optimizacion.precio_tablero:,.0f}"
            row += 1
            
            if optimizacion.num_tableros:
                costo_material = optimizacion.precio_tablero * optimizacion.num_tableros
                ws1[f'A{row}'] = f'Costo de Material ({optimizacion.num_tableros} tableros):'
                ws1[f'A{row}'].font = Font(bold=True)
                ws1[f'B{row}'] = f"${costo_material:,.0f}"
                row += 1
        
        if optimizacion.mano_obra:
            ws1[f'A{row}'] = 'Mano de Obra:'
            ws1[f'A{row}'].font = Font(bold=True)
            ws1[f'B{row}'] = f"${optimizacion.mano_obra:,.0f}"
            row += 1
        
        if costo_total:
            ws1[f'A{row}'] = 'COSTO TOTAL:'
            ws1[f'A{row}'].font = Font(bold=True, size=12)
            ws1[f'A{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            ws1[f'B{row}'] = f"${costo_total:,.0f}"
            ws1[f'B{row}'].font = Font(bold=True, size=12)
            ws1[f'B{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            row += 1
    
    # Ajustar ancho de columnas
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 30
    
    # === HOJA 2: LISTA DE PIEZAS ===
    ws2 = wb.create_sheet("Piezas")
    
    # Encabezados
    headers = ['Nombre', 'Ancho', 'Alto', 'Cantidad', 'Área Unitaria', 'Área Total']
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos de piezas
    unidad_opt = getattr(optimizacion, 'unidad_medida', 'cm') or 'cm'
    simbolo_area = obtener_simbolo_area(unidad_opt)
    factor_lineal = convertir_desde_cm(1, unidad_opt)
    factor_area = factor_lineal ** 2
    
    total_area_piezas = 0
    total_cantidad = 0
    
    for idx, pieza in enumerate(piezas_con_nombre, start=2):
        nombre = pieza.get('nombre', f"Pieza {idx-1}")
        ancho = float(pieza['ancho'])
        alto = float(pieza['alto'])
        cantidad = int(pieza['cantidad'])
        
        # Convertir a cm para cálculos
        ancho_cm = convertir_a_cm(ancho, unidad_opt)
        alto_cm = convertir_a_cm(alto, unidad_opt)
        area_unitaria_cm2 = ancho_cm * alto_cm
        area_total_cm2 = area_unitaria_cm2 * cantidad
        
        # Convertir para mostrar
        area_unitaria_mostrar = round(area_unitaria_cm2 * factor_area, 2)
        area_total_mostrar = round(area_total_cm2 * factor_area, 2)
        
        ws2.cell(row=idx, column=1, value=nombre).border = border
        ws2.cell(row=idx, column=2, value=ancho).border = border
        ws2.cell(row=idx, column=3, value=alto).border = border
        ws2.cell(row=idx, column=4, value=cantidad).border = border
        ws2.cell(row=idx, column=5, value=f"{area_unitaria_mostrar} {simbolo_area}").border = border
        ws2.cell(row=idx, column=6, value=f"{area_total_mostrar} {simbolo_area}").border = border
        
        total_area_piezas += area_total_cm2
        total_cantidad += cantidad
    
    # Fila de totales
    row_total = len(piezas_con_nombre) + 2
    ws2.cell(row=row_total, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=row_total, column=4, value=total_cantidad).font = Font(bold=True)
    ws2.cell(row=row_total, column=6, value=f"{round(total_area_piezas * factor_area, 2)} {simbolo_area}").font = Font(bold=True)
    
    # Ajustar ancho de columnas
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    
    # === HOJA 3: RESUMEN POR TABLERO ===
    ws3 = wb.create_sheet("Resumen por Tablero")
    
    # Estilos adicionales para desperdicio
    good_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # Verde claro
    warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Amarillo
    bad_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # Rojo claro
    
    # Encabezados
    headers = ['Tablero', 'Piezas', 'Área Disponible', 'Área Usada', 'Desperdicio', '% Desperdicio', '% Uso']
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos de tableros
    info_tableros = info_desperdicio.get('info_tableros', [])
    for idx, info in enumerate(info_tableros, start=2):
        area_disponible = info.get('area_usada', 0) + info.get('desperdicio', 0)
        porcentaje_desperdicio = (info.get('desperdicio', 0) / area_disponible * 100) if area_disponible > 0 else 0
        porcentaje_uso = info.get('porcentaje_uso', 0)
        
        ws3.cell(row=idx, column=1, value=f"Tablero {info.get('numero', idx-1)}").border = border
        ws3.cell(row=idx, column=1).alignment = center_alignment
        
        ws3.cell(row=idx, column=2, value=info.get('num_piezas', 0)).border = border
        ws3.cell(row=idx, column=2).alignment = center_alignment
        
        ws3.cell(row=idx, column=3, value=f"{area_disponible:.2f} {simbolo_area}").border = border
        ws3.cell(row=idx, column=3).alignment = center_alignment
        
        ws3.cell(row=idx, column=4, value=f"{info.get('area_usada', 0):.2f} {simbolo_area}").border = border
        ws3.cell(row=idx, column=4).alignment = center_alignment
        
        ws3.cell(row=idx, column=5, value=f"{info.get('desperdicio', 0):.2f} {simbolo_area}").border = border
        ws3.cell(row=idx, column=5).alignment = center_alignment
        
        # Columna % Desperdicio con colores
        cell_desperdicio = ws3.cell(row=idx, column=6, value=f"{porcentaje_desperdicio:.2f}%")
        cell_desperdicio.border = border
        cell_desperdicio.alignment = center_alignment
        if porcentaje_desperdicio < 20:
            cell_desperdicio.fill = good_fill
        elif porcentaje_desperdicio < 40:
            cell_desperdicio.fill = warning_fill
        else:
            cell_desperdicio.fill = bad_fill
        
        # Columna % Uso con colores
        cell_uso = ws3.cell(row=idx, column=7, value=f"{porcentaje_uso:.2f}%")
        cell_uso.border = border
        cell_uso.alignment = center_alignment
        if porcentaje_uso >= 80:
            cell_uso.fill = good_fill
        elif porcentaje_uso >= 60:
            cell_uso.fill = warning_fill
        else:
            cell_uso.fill = bad_fill
    
    # Fila de totales
    row_total = len(info_tableros) + 2
    ws3.cell(row=row_total, column=1, value="TOTALES").font = Font(bold=True, size=12)
    ws3.cell(row=row_total, column=1).border = border
    ws3.cell(row=row_total, column=1).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    total_area_disponible = sum(info.get('area_usada', 0) + info.get('desperdicio', 0) for info in info_tableros)
    total_area_usada = sum(info.get('area_usada', 0) for info in info_tableros)
    total_desperdicio = sum(info.get('desperdicio', 0) for info in info_tableros)
    total_porcentaje_desperdicio = (total_desperdicio / total_area_disponible * 100) if total_area_disponible > 0 else 0
    promedio_uso = sum(info.get('porcentaje_uso', 0) for info in info_tableros) / len(info_tableros) if info_tableros else 0
    
    ws3.cell(row=row_total, column=2, value=sum(info.get('num_piezas', 0) for info in info_tableros)).font = Font(bold=True)
    ws3.cell(row=row_total, column=2).border = border
    ws3.cell(row=row_total, column=2).alignment = center_alignment
    ws3.cell(row=row_total, column=2).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    ws3.cell(row=row_total, column=3, value=f"{total_area_disponible:.2f} {simbolo_area}").font = Font(bold=True)
    ws3.cell(row=row_total, column=3).border = border
    ws3.cell(row=row_total, column=3).alignment = center_alignment
    ws3.cell(row=row_total, column=3).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    ws3.cell(row=row_total, column=4, value=f"{total_area_usada:.2f} {simbolo_area}").font = Font(bold=True)
    ws3.cell(row=row_total, column=4).border = border
    ws3.cell(row=row_total, column=4).alignment = center_alignment
    ws3.cell(row=row_total, column=4).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    ws3.cell(row=row_total, column=5, value=f"{total_desperdicio:.2f} {simbolo_area}").font = Font(bold=True)
    ws3.cell(row=row_total, column=5).border = border
    ws3.cell(row=row_total, column=5).alignment = center_alignment
    ws3.cell(row=row_total, column=5).fill = bad_fill
    
    cell_total_desperdicio_pct = ws3.cell(row=row_total, column=6, value=f"{total_porcentaje_desperdicio:.2f}%")
    cell_total_desperdicio_pct.font = Font(bold=True, size=12)
    cell_total_desperdicio_pct.border = border
    cell_total_desperdicio_pct.alignment = center_alignment
    if total_porcentaje_desperdicio < 20:
        cell_total_desperdicio_pct.fill = good_fill
    elif total_porcentaje_desperdicio < 40:
        cell_total_desperdicio_pct.fill = warning_fill
    else:
        cell_total_desperdicio_pct.fill = bad_fill
    
    cell_total_uso = ws3.cell(row=row_total, column=7, value=f"{promedio_uso:.2f}%")
    cell_total_uso.font = Font(bold=True, size=12)
    cell_total_uso.border = border
    cell_total_uso.alignment = center_alignment
    if promedio_uso >= 80:
        cell_total_uso.fill = good_fill
    elif promedio_uso >= 60:
        cell_total_uso.fill = warning_fill
    else:
        cell_total_uso.fill = bad_fill
    
    # Ajustar ancho de columnas
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 18
    ws3.column_dimensions['F'].width = 16
    ws3.column_dimensions['G'].width = 12

    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

def generar_excel_historial_costos(optimizaciones_con_costo, estadisticas, fecha_desde=None, fecha_hasta=None):
    """
    Genera un archivo Excel con el resumen completo del historial de costos.
    
    Args:
        optimizaciones_con_costo: Lista de diccionarios con optimizaciones y costos
        estadisticas: Diccionario con estadísticas calculadas
        fecha_desde: Fecha desde (opcional)
        fecha_hasta: Fecha hasta (opcional)
    
    Returns:
        BytesIO: Buffer con el archivo Excel
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # === HOJA 1: RESUMEN DE ESTADÍSTICAS ===
    ws1 = wb.active
    ws1.title = "Resumen"
    
    # Título
    ws1['A1'] = "💰 HISTORIAL DE COSTOS - RESUMEN"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:B1')
    
    # Período
    row = 3
    if fecha_desde or fecha_hasta:
        periodo_texto = "Período: "
        if fecha_desde:
            periodo_texto += f"Desde {fecha_desde}"
        if fecha_desde and fecha_hasta:
            periodo_texto += " hasta "
        if fecha_hasta:
            periodo_texto += f"{fecha_hasta}"
        ws1[f'A{row}'] = periodo_texto
        ws1[f'A{row}'].font = Font(bold=True, italic=True)
        ws1.merge_cells(f'A{row}:B{row}')
        row += 2
    else:
        ws1[f'A{row}'] = "Período: Todos los registros"
        ws1[f'A{row}'].font = Font(bold=True, italic=True)
        ws1.merge_cells(f'A{row}:B{row}')
        row += 2
    
    # Estadísticas principales
    stats_data = [
        ['Total Optimizaciones:', estadisticas.get('total_optimizaciones', 0)],
        ['Costo Total:', f"${estadisticas.get('costo_total', 0):,.0f}"],
        ['Costo Material:', f"${estadisticas.get('costo_material_total', 0):,.0f}"],
        ['Mano de Obra:', f"${estadisticas.get('costo_mano_obra_total', 0):,.0f}"],
        ['Total Tableros:', estadisticas.get('num_tableros_total', 0)],
        ['Costo Promedio:', f"${estadisticas.get('costo_promedio', 0):,.0f}"],
        ['Costo por Tablero:', f"${estadisticas.get('costo_por_tablero', 0):,.0f}"],
        ['% Material:', f"{estadisticas.get('porcentaje_material', 0):.1f}%"],
        ['% Mano de Obra:', f"{estadisticas.get('porcentaje_mano_obra', 0):.1f}%"],
        ['Aprovechamiento Promedio:', f"{estadisticas.get('aprovechamiento_promedio', 0):.1f}%"],
    ]
    
    for label, valor in stats_data:
        ws1[f'A{row}'] = label
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = valor
        if 'Costo Total' in label:
            ws1[f'B{row}'].font = Font(bold=True, size=12, color="006100")
        row += 1
    
    # Ajustar ancho de columnas
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20
    
    # === HOJA 2: DETALLE DE OPTIMIZACIONES ===
    ws2 = wb.create_sheet("Detalle de Costos")
    
    # Encabezados
    headers = ['Fecha', 'Hora', 'Dimensiones', 'Material', 'Tableros', 'Aprovechamiento', 
               'Costo Material', 'Mano de Obra', 'Costo Total']
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos de optimizaciones
    for idx, item in enumerate(optimizaciones_con_costo, start=2):
        opt = item['optimizacion']
        
        # Fecha y hora
        fecha_str = opt.fecha.strftime('%d/%m/%Y')
        hora_str = opt.fecha.strftime('%H:%M')
        
        # Dimensiones
        dim_str = f"{opt.ancho_tablero} × {opt.alto_tablero} cm"
        if opt.unidad_medida != 'cm':
            dim_str += f" ({opt.get_unidad_medida_display()})"
        
        # Material
        material_str = opt.material.nombre if opt.material else "-"
        
        # Tableros
        num_tableros = opt.num_tableros or 0
        
        # Aprovechamiento
        aprovechamiento = opt.aprovechamiento_total or 0
        
        # Costos
        costo_material = item['costo_material']
        costo_mano_obra = item['costo_mano_obra']
        costo_total = item['costo_total']
        
        # Escribir datos
        ws2.cell(row=idx, column=1, value=fecha_str).border = border
        ws2.cell(row=idx, column=2, value=hora_str).border = border
        ws2.cell(row=idx, column=3, value=dim_str).border = border
        ws2.cell(row=idx, column=4, value=material_str).border = border
        ws2.cell(row=idx, column=5, value=num_tableros).border = border
        ws2.cell(row=idx, column=6, value=f"{aprovechamiento:.1f}%").border = border
        ws2.cell(row=idx, column=7, value=f"${costo_material:,.0f}").border = border
        ws2.cell(row=idx, column=8, value=f"${costo_mano_obra:,.0f}").border = border
        ws2.cell(row=idx, column=9, value=f"${costo_total:,.0f}").border = border
        
        # Formato de números
        ws2.cell(row=idx, column=5).alignment = center_alignment
        ws2.cell(row=idx, column=6).alignment = center_alignment
        ws2.cell(row=idx, column=7).alignment = Alignment(horizontal='right')
        ws2.cell(row=idx, column=8).alignment = Alignment(horizontal='right')
        ws2.cell(row=idx, column=9).alignment = Alignment(horizontal='right')
        ws2.cell(row=idx, column=9).font = Font(bold=True)
    
    # Fila de totales
    row_total = len(optimizaciones_con_costo) + 2
    ws2.cell(row=row_total, column=1, value="TOTALES:").font = Font(bold=True, size=12)
    ws2.cell(row=row_total, column=1).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    ws2.merge_cells(f'A{row_total}:F{row_total}')
    
    ws2.cell(row=row_total, column=7, value=f"${estadisticas.get('costo_material_total', 0):,.0f}").font = Font(bold=True)
    ws2.cell(row=row_total, column=7).border = border
    ws2.cell(row=row_total, column=7).alignment = Alignment(horizontal='right')
    ws2.cell(row=row_total, column=7).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    ws2.cell(row=row_total, column=8, value=f"${estadisticas.get('costo_mano_obra_total', 0):,.0f}").font = Font(bold=True)
    ws2.cell(row=row_total, column=8).border = border
    ws2.cell(row=row_total, column=8).alignment = Alignment(horizontal='right')
    ws2.cell(row=row_total, column=8).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    ws2.cell(row=row_total, column=9, value=f"${estadisticas.get('costo_total', 0):,.0f}").font = Font(bold=True, size=12, color="006100")
    ws2.cell(row=row_total, column=9).border = border
    ws2.cell(row=row_total, column=9).alignment = Alignment(horizontal='right')
    ws2.cell(row=row_total, column=9).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    # Ajustar ancho de columnas
    column_widths = [12, 8, 18, 20, 10, 15, 15, 15, 15]
    for col, width in enumerate(column_widths, start=1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

def generar_excel_resumen_desperdicio(optimizaciones, estadisticas, periodo='todos'):
    """
    Genera un archivo Excel con el resumen de desperdicio desde estadísticas.
    
    Args:
        optimizaciones: QuerySet de optimizaciones filtradas
        estadisticas: Diccionario con estadísticas calculadas
        periodo: Período seleccionado (todos, semanal, mensual, anual)
    
    Returns:
        BytesIO: Buffer con el archivo Excel
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # === HOJA 1: RESUMEN DE DESPERDICIO ===
    ws1 = wb.active
    ws1.title = "Resumen Desperdicio"
    
    # Título
    ws1['A1'] = "📊 RESUMEN DE DESPERDICIO - REPORTE DE EFICIENCIA"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:C1')
    
    # Período
    row = 3
    periodo_texto = {
        'todos': 'Todos los tiempos',
        'semanal': 'Última semana',
        'mensual': 'Último mes',
        'anual': 'Último año'
    }.get(periodo, 'Todos los tiempos')
    ws1[f'A{row}'] = f"Período: {periodo_texto}"
    ws1[f'A{row}'].font = Font(bold=True, italic=True)
    ws1.merge_cells(f'A{row}:C{row}')
    row += 2
    
    # Estadísticas principales
    stats_data = [
        ['Total Optimizaciones:', estadisticas.get('total_optimizaciones', 0)],
        ['Aprovechamiento Promedio:', f"{estadisticas.get('promedio_aprovechamiento', 0):.2f}%"],
        ['Desperdicio Promedio:', f"{estadisticas.get('promedio_desperdicio', 0):.2f}%"],
        ['Mejor Aprovechamiento:', f"{estadisticas.get('max_aprovechamiento', 0):.2f}%"],
        ['Peor Aprovechamiento:', f"{estadisticas.get('min_aprovechamiento', 0):.2f}%"],
    ]
    
    for label, valor in stats_data:
        ws1[f'A{row}'] = label
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = valor
        if 'Desperdicio' in label:
            ws1[f'B{row}'].font = Font(bold=True, size=12, color="DC3545")
        row += 1
    
    # Ajustar ancho de columnas
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20
    
    # === HOJA 2: DETALLE DE OPTIMIZACIONES ===
    ws2 = wb.create_sheet("Detalle Optimizaciones")
    
    # Encabezados
    headers = ['Fecha', 'Hora', 'Dimensiones', 'Aprovechamiento', 'Desperdicio', 
               'Tableros', 'Área Total', 'Área Usada', 'Área Desperdiciada']
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos de optimizaciones
    for idx, opt in enumerate(optimizaciones.order_by('fecha'), start=2):
        # Fecha y hora
        fecha_str = opt.fecha.strftime('%d/%m/%Y')
        hora_str = opt.fecha.strftime('%H:%M')
        
        # Dimensiones
        unidad_opt = getattr(opt, 'unidad_medida', 'cm') or 'cm'
        ancho_mostrar = round(convertir_desde_cm(opt.ancho_tablero, unidad_opt), 2)
        alto_mostrar = round(convertir_desde_cm(opt.alto_tablero, unidad_opt), 2)
        simbolo = obtener_simbolo_unidad(unidad_opt)
        dim_str = f"{ancho_mostrar} × {alto_mostrar} {simbolo}"
        
        # Aprovechamiento y desperdicio
        aprovechamiento = opt.aprovechamiento_total or 0
        desperdicio = 100 - aprovechamiento
        
        # Tableros
        num_tableros = opt.num_tableros or 1
        
        # Calcular áreas (necesitamos regenerar el gráfico para obtener info_desperdicio)
        # Por ahora usamos valores aproximados basados en aprovechamiento
        area_tablero = opt.ancho_tablero * opt.alto_tablero  # en cm²
        area_total = area_tablero * num_tableros
        area_usada = area_total * (aprovechamiento / 100)
        area_desperdiciada = area_total - area_usada
        
        simbolo_area = obtener_simbolo_area(unidad_opt)
        
        # Escribir datos
        ws2.cell(row=idx, column=1, value=fecha_str).border = border
        ws2.cell(row=idx, column=2, value=hora_str).border = border
        ws2.cell(row=idx, column=3, value=dim_str).border = border
        ws2.cell(row=idx, column=4, value=f"{aprovechamiento:.2f}%").border = border
        ws2.cell(row=idx, column=5, value=f"{desperdicio:.2f}%").border = border
        ws2.cell(row=idx, column=6, value=num_tableros).border = border
        ws2.cell(row=idx, column=7, value=f"{area_total:.2f} {simbolo_area}").border = border
        ws2.cell(row=idx, column=8, value=f"{area_usada:.2f} {simbolo_area}").border = border
        ws2.cell(row=idx, column=9, value=f"{area_desperdiciada:.2f} {simbolo_area}").border = border
        
        # Formato de números
        ws2.cell(row=idx, column=4).alignment = center_alignment
        ws2.cell(row=idx, column=5).alignment = center_alignment
        ws2.cell(row=idx, column=6).alignment = center_alignment
        
        # Colorear según desperdicio
        if desperdicio > 30:
            fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        elif desperdicio > 15:
            fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        else:
            fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        ws2.cell(row=idx, column=5).fill = fill
    
    # Fila de totales
    row_total = len(optimizaciones) + 2
    ws2.cell(row=row_total, column=1, value="TOTALES:").font = Font(bold=True, size=12)
    ws2.cell(row=row_total, column=1).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    ws2.merge_cells(f'A{row_total}:C{row_total}')
    
    # Ajustar ancho de columnas
    column_widths = [12, 8, 18, 15, 15, 10, 15, 15, 18]
    for col, width in enumerate(column_widths, start=1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
