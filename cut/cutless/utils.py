# Imports estándar de Python
import io
import os
import base64
from collections import Counter
from decimal import Decimal

# Imports de terceros
import matplotlib
matplotlib.use('Agg')
import matplotlib.dates as mdates
import matplotlib.patches as patches
import matplotlib.pyplot as plt
from PIL import Image
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

# Imports de Django
from django.conf import settings


# ===== FUNCIONES DE CONVERSIÓN DE UNIDADES =====
# Todas las conversiones son a centímetros (unidad base)

def convertir_a_cm(valor, unidad):
    """
    Convierte un valor de cualquier unidad a centímetros.
    
    Args:
        valor: Valor numérico a convertir
        unidad: Unidad de origen ('cm', 'm', 'mm', 'in', 'ft', 'pulgadas')
    
    Returns:
        Valor convertido a centímetros
    """
    # Normalizar 'pulgadas' a 'in' para compatibilidad
    if unidad == 'pulgadas':
        unidad = 'in'
    
    conversiones = {
        'cm': 1.0,           # 1 cm = 1 cm
        'm': 100.0,          # 1 m = 100 cm
        'mm': 0.1,           # 1 mm = 0.1 cm
        'in': 2.54,          # 1 pulgada = 2.54 cm
        'ft': 30.48,         # 1 pie = 30.48 cm
    }
    factor = conversiones.get(unidad, 1.0)
    return round(valor * factor, 2)


def convertir_desde_cm(valor_cm, unidad_destino):
    """
    Convierte un valor en centímetros a otra unidad.
    
    Args:
        valor_cm: Valor en centímetros
        unidad_destino: Unidad de destino ('cm', 'm', 'mm', 'in', 'ft', 'pulgadas')
    
    Returns:
        Valor convertido a la unidad destino
    """
    # Normalizar 'pulgadas' a 'in' para compatibilidad
    if unidad_destino == 'pulgadas':
        unidad_destino = 'in'
    
    conversiones = {
        'cm': 1.0,           # 1 cm = 1 cm
        'm': 0.01,           # 1 cm = 0.01 m
        'mm': 10.0,          # 1 cm = 10 mm
        'in': 1/2.54,        # 1 cm = 1/2.54 pulgadas
        'ft': 1/30.48,       # 1 cm = 1/30.48 pies
    }
    factor = conversiones.get(unidad_destino, 1.0)
    return round(valor_cm * factor, 2)


def obtener_simbolo_unidad(unidad):
    """
    Retorna el símbolo de la unidad para mostrar.
    """
    # Normalizar 'pulgadas' a 'in' para compatibilidad
    if unidad == 'pulgadas':
        unidad = 'in'
    
    simbolos = {
        'cm': 'cm',
        'm': 'm',
        'mm': 'mm',
        'in': 'in',
        'ft': 'ft',
    }
    return simbolos.get(unidad, 'cm')


def obtener_simbolo_area(unidad):
    """
    Retorna el símbolo de unidad al cuadrado para áreas.
    """
    # Normalizar 'pulgadas' a 'in' para compatibilidad
    if unidad == 'pulgadas':
        unidad = 'in'
    
    simbolos = {
        'cm': 'cm²',
        'm': 'm²',
        'mm': 'mm²',
        'in': 'in²',
        'ft': 'ft²',
    }
    return simbolos.get(unidad, 'cm²')


def pieza_cabe_en_tablero(pieza_ancho_cm, pieza_alto_cm, tablero_ancho_cm, tablero_alto_cm, permitir_rotacion=True):
    """
    True si la pieza puede cortarse en una sola placa sin partirla
    (orientación original u orientación rotada 90° si permitir_rotacion).
    """
    w = float(pieza_ancho_cm)
    h = float(pieza_alto_cm)
    W = float(tablero_ancho_cm)
    H = float(tablero_alto_cm)
    if w <= W and h <= H:
        return True
    if permitir_rotacion and h <= W and w <= H:
        return True
    return False


def _kern_block(w, h, x, y, W, H, m):
    """
    Hueco ocupado tras colocar pieza geométrica w×h en (x,y): añade kerf solo
    si no tocamos borde derecho/inferior del tablero.
    """
    bw = float(w + (m if x + float(w) < W - 1e-9 else 0))
    bh = float(h + (m if y + float(h) < H - 1e-9 else 0))
    return bw, bh


def _subtract_rect(ax, ay, aw, ah, bx, by, bw, bh, eps=1e-9):
    """
    Parte el rectángulo A\(A∩B) en hasta 4 rectángulos (ancla arriba-izquierda, eje y hacia abajo).
    Devuelve lista de tuplas (x, y, w, h).
    """
    ix0 = max(ax, bx)
    iy0 = max(ay, by)
    ix1 = min(ax + aw, bx + bw)
    iy1 = min(ay + ah, by + bh)
    if ix0 >= ix1 - eps or iy0 >= iy1 - eps:
        return [(ax, ay, aw, ah)]
    chunks = []
    if iy0 > ay + eps:
        chunks.append((ax, ay, aw, iy0 - ay))
    if ay + ah > iy1 + eps:
        chunks.append((ax, iy1, aw, ay + ah - iy1))
    mh = iy1 - iy0
    if ix0 > ax + eps:
        chunks.append((ax, iy0, ix0 - ax, mh))
    if ax + aw > ix1 + eps:
        chunks.append((ix1, iy0, ax + aw - ix1, mh))
    return [(x, y, rw, rh) for x, y, rw, rh in chunks if rw > eps and rh > eps]


def _merge_free_rects(rects, eps=1e-9):
    """Elimina rectángulos totalmente contenidos en otro."""
    rects = [(float(x), float(y), float(w), float(h)) for x, y, w, h in rects if w > eps and h > eps]
    kept = []
    n = len(rects)
    for i in range(n):
        x, y, w, h = rects[i]
        x2, y2 = x + w, y + h
        inside_other = False
        for j in range(n):
            if i == j:
                continue
            ox, oy, ow, oh = rects[j]
            if (
                ox - eps <= x
                and oy - eps <= y
                and ox + ow + eps >= x2
                and oy + oh + eps >= y2
            ):
                inside_other = True
                break
        if not inside_other:
            kept.append((x, y, w, h))
    return kept


def _fuse_adjacent_free_rects(rects, eps=1e-5):
    """
    Une rectángulos libres del mismo ancho y apilados en vertical (o misma altura,
    recorridos en horizontal) para que el empaquetador “vea” columnas y bandas enteras
    y no solo franjas cortadas por colocaciones anteriores.
    """
    rects = [(float(x), float(y), float(w), float(h)) for x, y, w, h in rects if w > eps and h > eps]
    changed = True
    while changed and len(rects) > 1:
        changed = False
        n = len(rects)
        for i in range(n):
            for j in range(i + 1, n):
                ax, ay, aw, ah = rects[i]
                bx, by, bw, bh = rects[j]
                if abs(ax - bx) < eps and abs(aw - bw) < eps:
                    if abs((ay + ah) - by) < eps or abs((by + bh) - ay) < eps:
                        y0 = min(ay, by)
                        y1 = max(ay + ah, by + bh)
                        merged = (ax, y0, aw, y1 - y0)
                        rects = [merged] + [rects[k] for k in range(n) if k not in (i, j)]
                        changed = True
                        break
                if abs(ay - by) < eps and abs(ah - bh) < eps:
                    if abs((ax + aw) - bx) < eps or abs((bx + bw) - ax) < eps:
                        x0 = min(ax, bx)
                        x1 = max(ax + aw, bx + bw)
                        merged = (x0, ay, x1 - x0, ah)
                        rects = [merged] + [rects[k] for k in range(n) if k not in (i, j)]
                        changed = True
                        break
            if changed:
                break
    return rects


def _normalizar_rects_libres(rects, eps=1e-5):
    """Contención + fusión de adyacentes hasta estabilizar."""
    r = _merge_free_rects(rects)
    prev = None
    while prev != r:
        prev = r
        r = _fuse_adjacent_free_rects(r, eps)
        r = _merge_free_rects(r)
    return r


def mensaje_advertencia_piezas_no_colocadas(info_desperdicio, unidad='cm'):
    """
    Mensaje para el usuario cuando hubo piezas omitidas por no caber en el tablero.
    Devuelve None si no hay piezas omitidas.
    """
    omitidas = (info_desperdicio or {}).get('piezas_no_colocadas') or []
    if not omitidas:
        return None

    solicitadas = (info_desperdicio or {}).get('num_piezas_solicitadas')
    colocadas = (info_desperdicio or {}).get('num_piezas_colocadas')
    simbolo = obtener_simbolo_unidad(unidad)
    cuenta = Counter()
    for p in omitidas:
        key = (p.get('nombre', 'Pieza'), float(p.get('ancho_cm', 0)), float(p.get('alto_cm', 0)))
        cuenta[key] += 1
    partes_desc = []
    for (nombre, aw, ah), cnt in cuenta.items():
        wu = round(convertir_desde_cm(aw, unidad), 2)
        hu = round(convertir_desde_cm(ah, unidad), 2)
        partes_desc.append(f"{nombre} ×{cnt} ({wu}×{hu} {simbolo})")
    partes_desc.sort()
    max_items = 12
    texto = "; ".join(partes_desc[:max_items])
    if len(partes_desc) > max_items:
        texto += f"; …(+{len(partes_desc) - max_items})"
    sufix = ""
    if solicitadas is not None and colocadas is not None:
        sufix = f" Piezas en el plano: {colocadas} de {solicitadas}."
    elif colocadas is not None:
        sufix = f" Piezas en el plano: {colocadas}."
    return (
        f"No se pudieron colocar {len(omitidas)} pieza(s) (dimensiones mayores que el tablero en una sola orientación válida): {texto}.{sufix}"
    )


def generar_grafico(piezas, ancho_tablero, alto_tablero, unidad='cm', permitir_rotacion=True, margen_corte=0.3, nombres_piezas=None, modo_plan_corte=False):
    """
    Piezas ordenadas por área (FFD) + colocación en rectángulos libres (BSSF): no todas
    las piezas comparten una misma “fila” rígida; se pueden colocar en huecos bajo más altas
    cuando la geometría y el kerf lo permiten. Rotación opcional entre orientaciones válidas.
    
    Args:
        piezas: Lista de tuplas (ancho, alto, cantidad) en cm
        ancho_tablero: Ancho del tablero en cm
        alto_tablero: Alto del tablero en cm
        unidad: Unidad de medida para mostrar
        permitir_rotacion: Si True, intenta rotar piezas 90° si mejora el aprovechamiento
        margen_corte: Margen de corte (kerf) en cm entre cortes vecinos (no sobre borde placa).
        nombres_piezas: Lista opcional de nombres para la etiqueta en el gráfico
    """
    AREA_TABLERO = ancho_tablero * alto_tablero
    tableros = []
    area_usada_total = 0
    piezas_no_colocadas = []
    num_piezas_solicitadas = sum(int(c) for _, _, c in piezas)

    # Paso 1: Expandir piezas y ordenarlas (FFD)
    piezas_expandidas = []
    for idx, (w, h, c) in enumerate(piezas):
        nombre_base = None
        if nombres_piezas and idx < len(nombres_piezas):
            nombre_base = nombres_piezas[idx].strip() or f"Pieza {idx+1}"
        else:
            nombre_base = f"Pieza {idx+1}"

        for _ in range(c):
            area = w * h
            piezas_expandidas.append({
                'ancho': w,
                'alto': h,
                'area': area,
                'original': (w, h),
                'rotada': False,
                'nombre': nombre_base
            })
    
    piezas_expandidas.sort(key=lambda x: x['area'], reverse=True)
    
    EPS = 1e-9
    W_bin = float(ancho_tablero)
    H_bin = float(alto_tablero)
    _m = float(margen_corte)

    def _calcular_desperdicio(tablero):
        area_usada = 0
        for pos in tablero['posiciones']:
            if len(pos) >= 5:
                _, _, wg, hg, _ = pos[:5]
                area_usada += wg * hg
        return AREA_TABLERO - area_usada

    def _tb_clonar(tb):
        return {
            'posiciones': list(tb['posiciones']),
            'free_rects': [(float(a), float(b), float(fw), float(fh)) for a, b, fw, fh in tb['free_rects']],
        }

    def _tb_vacio():
        return {'posiciones': [], 'free_rects': [(0.0, 0.0, W_bin, H_bin)]}

    def _mejor_ancla_bssf(tablero, wg, hg):
        """Esquina sup-izquierda del hueco: BSSF dentro de rectángulos libres."""
        best_key = None
        anchor = None
        for fx, fy, fw, fh in tablero['free_rects']:
            if fx + wg > W_bin + EPS or fy + hg > H_bin + EPS:
                continue
            bw, bh = _kern_block(wg, hg, fx, fy, W_bin, H_bin, _m)
            if bw > fw + EPS or bh > fh + EPS:
                continue
            ss = fw - bw
            ls = fh - bh
            key = (ss, ls, fy, fx)
            if best_key is None or key < best_key:
                best_key = key
                anchor = (fx, fy, bw, bh)
        return anchor

    def _aplicar_pieza(tablero, gx, gy, bw, bh, wg, hg, rotada, wo, ho, nombre):
        tablero['posiciones'].append((gx, gy, wg, hg, rotada, wo, ho, nombre))
        nueva = []
        for fx, fy, fw, fh in tablero['free_rects']:
            nueva.extend(_subtract_rect(fx, fy, fw, fh, gx, gy, bw, bh))
        tablero['free_rects'] = _normalizar_rects_libres(nueva)

    def _intentar_colocacion(tb_orig, pieza, wg, hg, rotada):
        cand = _tb_clonar(tb_orig)
        pack = _mejor_ancla_bssf(cand, wg, hg)
        if pack is None:
            return None
        gx, gy, bw, bh = pack
        _aplicar_pieza(
            cand,
            gx,
            gy,
            bw,
            bh,
            wg,
            hg,
            rotada,
            pieza['ancho'],
            pieza['alto'],
            pieza.get('nombre', 'Pieza'),
        )
        return cand

    def _considerar(best, nuevo_key, payload):
        if best is None or nuevo_key < best[0]:
            return (nuevo_key, payload)
        return best

    # Paso 2: mismo orden FFD pero colocación en rectángulos libres
    for pieza in piezas_expandidas:
        w_original = pieza['ancho']
        h_original = pieza['alto']
        if permitir_rotacion and w_original != h_original:
            orientaciones = [(w_original, h_original, False), (h_original, w_original, True)]
        else:
            orientaciones = [(w_original, h_original, False)]

        mejor = None  # (clave de orden lex, payload)

        for wg, hg, rot in orientaciones:
            if wg > W_bin + EPS or hg > H_bin + EPS:
                continue

            for tbi, tb in enumerate(tableros):
                prob = _intentar_colocacion(tb, pieza, wg, hg, rot)
                if prob is None:
                    continue
                desp = _calcular_desperdicio(prob)
                k = (desp, 1 if rot else 0, tbi)
                mejor = _considerar(mejor, k, ('existe', tbi, prob, rot))

            nueva_hoja = _intentar_colocacion(_tb_vacio(), pieza, wg, hg, rot)
            if nueva_hoja is None:
                continue
            desp_n = _calcular_desperdicio(nueva_hoja)
            # Sentinela: tableros existentes mejor que nueva lámina igual desperdicio
            k_n = (desp_n, 1 if rot else 0, len(tableros) + 1)
            mejor = _considerar(mejor, k_n, ('nuevo', nueva_hoja, rot))

        if mejor is None:
            piezas_no_colocadas.append({
                'nombre': pieza.get('nombre', 'Pieza'),
                'ancho_cm': w_original,
                'alto_cm': h_original,
            })
            continue

        _, payload = mejor
        if payload[0] == 'existe':
            _, tbi, nueva_tb, rot = payload
            tableros[tbi] = nueva_tb
            pieza['rotada'] = rot
        else:
            _, nueva_tb, rot = payload
            tableros.append(nueva_tb)
            pieza['rotada'] = rot
        area_usada_total += w_original * h_original

    # Paso 3: Calcular aprovechamiento y desperdicio
    num_tableros = len(tableros)
    num_piezas_colocadas = sum(len(t['posiciones']) for t in tableros)
    if num_tableros == 0:
        aprovechamiento_total = 0
        desperdicio_total = 0
        info_tableros = []
    else:
        area_total_disponible = num_tableros * AREA_TABLERO
        desperdicio_total = area_total_disponible - area_usada_total
        aprovechamiento_total = round((area_usada_total / area_total_disponible) * 100, 2)
        
        # Calcular desperdicio por tablero
        info_tableros = []
        for idx, tablero in enumerate(tableros, start=1):
            # Las posiciones ahora incluyen: (x, y, w, h, rotada, w_orig, h_orig) o (x, y, w, h, rotada)
            # Calcular área usando las dimensiones colocadas (w, h)
            area_usada_tablero = 0
            for pos in tablero['posiciones']:
                if len(pos) >= 5:
                    _, _, w, h, _ = pos[:5]  # Obtener w y h (dimensiones colocadas)
                    area_usada_tablero += w * h
            desperdicio_tablero = AREA_TABLERO - area_usada_tablero
            porcentaje_uso = round((area_usada_tablero / AREA_TABLERO) * 100, 2)
            
            info_tableros.append({
                'numero': idx,
                'area_usada': area_usada_tablero,
                'desperdicio': desperdicio_tablero,
                'porcentaje_uso': porcentaje_uso,
                'num_piezas': len(tablero['posiciones'])
            })
    
    # Paso 4: Generar imágenes
    imagenes_base64 = []
    
    for i, tablero in enumerate(tableros, start=1):
        posiciones = tablero['posiciones']
        info = info_tableros[i-1]
        
        fig, ax = plt.subplots(figsize=(10, 10))
        ax.set_xlim(0, ancho_tablero)
        ax.set_ylim(0, alto_tablero)
        ax.invert_yaxis()
        ax.set_aspect('equal')
        
        # Convertir desperdicio para mostrar
        simbolo = obtener_simbolo_unidad(unidad)
        simbolo_area = obtener_simbolo_area(unidad)
        factor_lineal = convertir_desde_cm(1, unidad)
        factor_area = factor_lineal ** 2
        desperdicio_mostrar = round(info['desperdicio'] * factor_area, 2)
        
        # Título y configuración según modo
        if modo_plan_corte:
            # Modo plan de corte: sin título, solo dimensiones del tablero
            ax.set_title(f"{ancho_tablero} × {alto_tablero} {simbolo}",
                        fontsize=12, fontweight='bold', pad=10)
            ax.set_xlabel(f"Ancho ({simbolo})", fontsize=10)
            ax.set_ylabel(f"Alto ({simbolo})", fontsize=10)
            ax.grid(True, alpha=0.2, linestyle='-', linewidth=0.5, color='gray')
        else:
            # Modo normal: con información completa
            ax.set_title(f"Tablero {i} de {num_tableros} - FFD + rect. libres\n"
                        f"Uso: {info['porcentaje_uso']}% | Desperdicio: {desperdicio_mostrar} {simbolo_area}",
                        fontsize=13, fontweight='bold', pad=20)
            ax.set_xlabel(f"Ancho ({simbolo})", fontsize=11)
            ax.set_ylabel(f"Alto ({simbolo})", fontsize=11)
            ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
        
        ax.set_axisbelow(True)
        
        # Borde del tablero
        if modo_plan_corte:
            borde = patches.Rectangle((0, 0), ancho_tablero, alto_tablero,
                                      linewidth=2, edgecolor='black', 
                                      facecolor='white', alpha=1.0)
        else:
            borde = patches.Rectangle((0, 0), ancho_tablero, alto_tablero,
                                      linewidth=3, edgecolor='black', 
                                      facecolor='#f0f0f0', alpha=0.3)
        ax.add_patch(borde)
        
        # Dibujar piezas
        if modo_plan_corte:
            # Modo plan de corte: blanco y negro, solo medidas y nombre de pieza
            for idx, pos_data in enumerate(posiciones):
                # Manejar diferentes formatos de posiciones
                if len(pos_data) >= 8:
                    x, y, w, h, rotada, w_orig, h_orig, nombre_pieza = pos_data
                elif len(pos_data) == 7:
                    x, y, w, h, rotada, w_orig, h_orig = pos_data
                    nombre_pieza = f'Pieza {idx+1}'
                elif len(pos_data) == 5:
                    x, y, w, h, rotada = pos_data
                    w_orig, h_orig = w, h
                    nombre_pieza = f'Pieza {idx+1}'
                else:
                    x, y, w, h = pos_data
                    rotada = False
                    w_orig, h_orig = w, h
                    nombre_pieza = f'Pieza {idx+1}'
                
                # Rectángulo blanco con borde negro
                rect = patches.Rectangle((x, y), w, h, linewidth=1.5,
                                         edgecolor='black', facecolor='white', alpha=1.0)
                ax.add_patch(rect)
                
                # Mostrar dimensiones en negro
                w_orig_mostrar = round(convertir_desde_cm(w_orig, unidad), 1)
                h_orig_mostrar = round(convertir_desde_cm(h_orig, unidad), 1)
                
                # Texto incluye nombre de la pieza + dimensiones
                texto_dimensiones = f'{nombre_pieza}\n{w_orig_mostrar}×{h_orig_mostrar}'
                if rotada:
                    texto_dimensiones += '\n(ROTADA 90°)'
                
                ax.text(x + w/2, y + h/2, texto_dimensiones,
                       ha='center', va='center', fontsize=7,
                       fontweight='bold', color='black',
                       bbox=dict(boxstyle='round,pad=0.2', facecolor='white',
                                edgecolor='black', linewidth=0.5, alpha=0.9))
        else:
            # Modo normal: colores y más información
            colores = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E2']
            
            for idx, pos_data in enumerate(posiciones):
                # Manejar diferentes formatos de posiciones
                if len(pos_data) >= 8:
                    x, y, w, h, rotada, w_orig, h_orig, nombre_pieza = pos_data
                elif len(pos_data) == 7:
                    x, y, w, h, rotada, w_orig, h_orig = pos_data
                    nombre_pieza = f'Pieza {idx+1}'
                elif len(pos_data) == 5:
                    x, y, w, h, rotada = pos_data
                    w_orig, h_orig = w, h
                    nombre_pieza = f'Pieza {idx+1}'
                else:
                    x, y, w, h = pos_data
                    rotada = False
                    w_orig, h_orig = w, h
                    nombre_pieza = f'Pieza {idx+1}'
                
                color = colores[idx % len(colores)]
                
                rect = patches.Rectangle((x, y), w, h, linewidth=2,
                                         edgecolor='darkblue', facecolor=color, alpha=0.7)
                ax.add_patch(rect)
                
                # Mostrar dimensiones
                w_orig_mostrar = round(convertir_desde_cm(w_orig, unidad), 1)
                h_orig_mostrar = round(convertir_desde_cm(h_orig, unidad), 1)
                
                # Mostrar dimensiones con indicador de rotación y nombre
                texto_dimensiones = f'{nombre_pieza}\n{w_orig_mostrar}×{h_orig_mostrar}'
                if rotada:
                    texto_dimensiones += '\n(ROTADA 90°)'
                
                ax.text(x + w/2, y + h/2, texto_dimensiones,
                       ha='center', va='center', fontsize=8, 
                       fontweight='bold', color='white',
                       bbox=dict(boxstyle='round,pad=0.3', facecolor='darkgreen' if rotada else 'black', alpha=0.8))
            
            # Información detallada (solo en modo normal)
            area_usada_mostrar = round(info['area_usada'] * factor_area, 2)
            info_text = (f"Piezas: {info['num_piezas']}\n"
                        f"Área usada: {area_usada_mostrar} {simbolo_area}\n"
                        f"Desperdicio: {desperdicio_mostrar} {simbolo_area}")
            ax.text(ancho_tablero * 0.02, alto_tablero * 0.98, info_text,
                   fontsize=9, verticalalignment='top',
                   bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.9))
        
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=120, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        buf.seek(0)
        imagenes_base64.append(base64.b64encode(buf.read()).decode("utf-8"))
    
    # Retornar también información de desperdicio
    return imagenes_base64, aprovechamiento_total, {
        'area_usada_total': area_usada_total,
        'desperdicio_total': desperdicio_total,
        'info_tableros': info_tableros,
        'num_tableros': num_tableros,
        'area_total_disponible': num_tableros * AREA_TABLERO if num_tableros > 0 else 0,
        'piezas_no_colocadas': piezas_no_colocadas,
        'num_piezas_solicitadas': num_piezas_solicitadas,
        'num_piezas_colocadas': num_piezas_colocadas,
    }


def _info_desperdicio_desde_optimizacion(optimizacion):
    """Regenera el dict info_desperdicio de generar_grafico desde datos guardados en BD."""
    unidad_opt = getattr(optimizacion, 'unidad_medida', 'cm') or 'cm'
    piezas = []
    nombres_piezas = []
    for linea in (optimizacion.piezas or '').splitlines():
        linea = linea.strip()
        if not linea:
            continue
        partes = [p.strip() for p in linea.split(',')]
        try:
            if len(partes) == 4:
                nombre, w, h, c_s = partes
                nombres_piezas.append(nombre)
                piezas.append((
                    convertir_a_cm(float(w), unidad_opt),
                    convertir_a_cm(float(h), unidad_opt),
                    int(c_s),
                ))
            elif len(partes) == 3:
                w, h, c_s = partes
                nombres_piezas.append(f"Pieza {len(nombres_piezas) + 1}")
                piezas.append((
                    convertir_a_cm(float(w), unidad_opt),
                    convertir_a_cm(float(h), unidad_opt),
                    int(c_s),
                ))
        except (ValueError, TypeError):
            continue
    if not piezas:
        return None
    permitir_rotacion = getattr(optimizacion, 'permitir_rotacion', True)
    margen_corte = getattr(optimizacion, 'margen_corte', 0.3) or 0.3
    _, _, info = generar_grafico(
        piezas,
        optimizacion.ancho_tablero,
        optimizacion.alto_tablero,
        unidad_opt,
        permitir_rotacion=permitir_rotacion,
        margen_corte=margen_corte,
        nombres_piezas=nombres_piezas if nombres_piezas else None,
    )
    return info


def generar_pdf(optimizacion, imagenes_base64, numero_lista=None, info_desperdicio=None):
    """
    Genera UN SOLO PDF con todos los tableros, cada uno en su propia página.
    
    Args:
        optimizacion: Objeto Optimizacion
        imagenes_base64: Lista de imágenes en base64
        numero_lista: Número de la lista en el historial (opcional). Si se proporciona,
                     se usará en el nombre del archivo en lugar del ID.
        info_desperdicio: Dict igual al tercer retorno de generar_grafico (areas en cm²).
                         Si es None, se regenera desde la optimización guardada (más costoso).
    """
    if isinstance(imagenes_base64, str):
        imagenes_base64 = [imagenes_base64] if imagenes_base64 else []
    
    if not imagenes_base64:
        return None
    
    # Usar número de lista si está disponible, sino usar el ID
    if numero_lista is not None:
        filename = f"optimizacion_{optimizacion.usuario.username}_{numero_lista}.pdf"
    else:
        filename = f"optimizacion_{optimizacion.usuario.username}_{optimizacion.id}.pdf"
    filepath = os.path.join(settings.MEDIA_ROOT, "pdfs", filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    c = canvas.Canvas(filepath, pagesize=A4)
    width, height = A4

    # === PÁGINA 1: Información general ===
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width / 2, height - 50, "Reporte de Optimización - CutLess")
    c.setFont("Helvetica-Oblique", 10)
    c.drawCentredString(width / 2, height - 70, "Algoritmo: First Fit Decreasing (FFD)")

    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, height - 110, "Información General:")
    c.setFont("Helvetica", 11)
    c.drawString(2.5*cm, height - 130, f"• Usuario: {optimizacion.usuario.username}")
    c.drawString(2.5*cm, height - 145, f"• Fecha: {optimizacion.fecha.strftime('%d/%m/%Y %H:%M')}")
    # Obtener unidad y convertir dimensiones para mostrar
    unidad = getattr(optimizacion, 'unidad_medida', 'cm') or 'cm'
    ancho_mostrar = convertir_desde_cm(optimizacion.ancho_tablero, unidad)
    alto_mostrar = convertir_desde_cm(optimizacion.alto_tablero, unidad)
    simbolo = obtener_simbolo_unidad(unidad)
    
    c.drawString(2.5*cm, height - 160, f"• Dimensiones del tablero: {ancho_mostrar} × {alto_mostrar} {simbolo}")
    c.drawString(2.5*cm, height - 175, f"• Tableros generados: {len(imagenes_base64)}")
    
    c.setFont("Helvetica-Bold", 12)
    c.setFillColorRGB(0, 0.5, 0)
    c.drawString(2.5*cm, height - 195, f"• Aprovechamiento: {optimizacion.aprovechamiento_total:.2f}%")
    c.setFillColorRGB(0, 0, 0)
    
    # Información de costos (si existe)
    costo_total = optimizacion.get_costo_total()
    if costo_total or optimizacion.precio_tablero or optimizacion.mano_obra:
        y_pos_costo = height - 220
        c.setFont("Helvetica-Bold", 12)
        c.setFillColorRGB(0, 0.5, 0.8)
        c.drawString(2*cm, y_pos_costo, "Información de Costos:")
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica", 11)
        y_pos_costo -= 20
        
        if optimizacion.material:
            c.drawString(2.5*cm, y_pos_costo, f"• Material: {optimizacion.material.nombre}")
            y_pos_costo -= 15
        
        if optimizacion.precio_tablero:
            c.drawString(2.5*cm, y_pos_costo, f"• Precio por tablero: ${optimizacion.precio_tablero:,.0f}")
            y_pos_costo -= 15
            if optimizacion.num_tableros:
                costo_material = optimizacion.precio_tablero * optimizacion.num_tableros
                c.drawString(2.5*cm, y_pos_costo, f"• Costo de material ({optimizacion.num_tableros} tableros): ${costo_material:,.0f}")
                y_pos_costo -= 15
        
        if optimizacion.mano_obra:
            c.drawString(2.5*cm, y_pos_costo, f"• Mano de obra: ${optimizacion.mano_obra:,.0f}")
            y_pos_costo -= 15
        
        if costo_total:
            c.setFont("Helvetica-Bold", 12)
            c.setFillColorRGB(0, 0.5, 0)
            c.drawString(2.5*cm, y_pos_costo, f"• COSTO TOTAL: ${costo_total:,.0f}")
            c.setFillColorRGB(0, 0, 0)
            y_pos_costo -= 20
        
        # Ajustar posición inicial de piezas
        height_piezas = y_pos_costo - 10
    else:
        height_piezas = height - 225

    # Encabezado de lista de piezas
    # Tabla de piezas detallada
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, height_piezas, "Lista de Piezas:")
    c.setFont("Helvetica-Bold", 10)
    
    # Encabezados de tabla
    y_pos = height_piezas - 20
    c.drawString(2.5*cm, y_pos, "Cant.")
    c.drawString(4*cm, y_pos, "Nombre")
    c.drawString(9*cm, y_pos, "Dimensiones")
    c.drawString(14*cm, y_pos, "Área unit.")
    c.drawString(17.5*cm, y_pos, "Área total")
    
    # Línea separadora
    y_pos -= 5
    c.line(2*cm, y_pos, width - 2*cm, y_pos)
    y_pos -= 15
    
    c.setFont("Helvetica", 9)
    total_area_piezas = 0
    total_cantidad_piezas = 0
    
    for linea in optimizacion.piezas.splitlines():
        if y_pos < 120:  # Más espacio para tabla
            c.showPage()
            y_pos = height - 50
            # Reimprimir encabezados si hay nueva página
            c.setFont("Helvetica-Bold", 10)
            c.drawString(2.5*cm, y_pos, "Cant.")
            c.drawString(4*cm, y_pos, "Nombre")
            c.drawString(9*cm, y_pos, "Dimensiones")
            c.drawString(14*cm, y_pos, "Área unit.")
            c.drawString(17.5*cm, y_pos, "Área total")
            y_pos -= 5
            c.line(2*cm, y_pos, width - 2*cm, y_pos)
            y_pos -= 15
            c.setFont("Helvetica", 9)
        
        try:
            partes = linea.split(',')
            if len(partes) == 4:
                nombre, ancho_str, alto_str, cantidad_str = partes
                nombre = nombre.strip()
                ancho = float(ancho_str.strip())
                alto = float(alto_str.strip())
                cantidad = int(cantidad_str.strip())
            else:
                ancho_str, alto_str, cantidad_str = partes
                nombre = "Pieza"
                ancho = float(ancho_str.strip())
                alto = float(alto_str.strip())
                cantidad = int(cantidad_str.strip())
            
            # Convertir a cm si es necesario
            ancho_cm = convertir_a_cm(ancho, optimizacion.unidad_medida)
            alto_cm = convertir_a_cm(alto, optimizacion.unidad_medida)
            
            # Calcular áreas
            area_unit_cm2 = ancho_cm * alto_cm
            area_total_cm2 = area_unit_cm2 * cantidad
            
            # Convertir para mostrar
            simbolo_area = obtener_simbolo_area(optimizacion.unidad_medida)
            factor_lineal = convertir_desde_cm(1, optimizacion.unidad_medida)
            factor_area = factor_lineal ** 2
            
            area_unit_mostrar = round(area_unit_cm2 * factor_area, 2)
            area_total_mostrar = round(area_total_cm2 * factor_area, 2)
            
            ancho_mostrar = round(convertir_desde_cm(ancho_cm, optimizacion.unidad_medida), 1)
            alto_mostrar = round(convertir_desde_cm(alto_cm, optimizacion.unidad_medida), 1)
            
            # Dibujar fila
            c.drawString(2.5*cm, y_pos, str(cantidad))
            c.drawString(4*cm, y_pos, nombre[:20])  # Limitar longitud
            c.drawString(9*cm, y_pos, f"{ancho_mostrar} × {alto_mostrar} {obtener_simbolo_unidad(optimizacion.unidad_medida)}")
            c.drawString(14*cm, y_pos, f"{area_unit_mostrar} {simbolo_area}")
            c.drawString(17.5*cm, y_pos, f"{area_total_mostrar} {simbolo_area}")
            
            total_area_piezas += area_total_cm2
            total_cantidad_piezas += cantidad
            
        except Exception as e:
            c.drawString(2.5*cm, y_pos, f"Error: {linea}")
        
        y_pos -= 15
    
    # Resumen de piezas
    y_pos -= 10
    c.line(2*cm, y_pos, width - 2*cm, y_pos)
    y_pos -= 15
    c.setFont("Helvetica-Bold", 10)
    simbolo_area = obtener_simbolo_area(optimizacion.unidad_medida)
    factor_lineal = convertir_desde_cm(1, optimizacion.unidad_medida)
    factor_area = factor_lineal ** 2
    total_area_mostrar = round(total_area_piezas * factor_area, 2)
    c.drawString(17.5*cm, y_pos, f"Total: {total_area_mostrar} {simbolo_area}")
    c.drawString(2.5*cm, y_pos, f"Total piezas: {total_cantidad_piezas}")
    
    # Información de configuración
    y_pos -= 30
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y_pos, "Configuración de optimización:")
    c.setFont("Helvetica", 9)
    y_pos -= 15
    rotacion_texto = "Sí" if getattr(optimizacion, 'permitir_rotacion', True) else "No"
    c.drawString(2.5*cm, y_pos, f"• Rotación automática: {rotacion_texto}")
    y_pos -= 12
    # El margen de corte se guarda en cm, pero siempre se muestra en mm
    margen_cm = getattr(optimizacion, 'margen_corte', 0.3)
    margen_mm = round(margen_cm * 10, 1)  # Convertir de cm a mm
    c.drawString(2.5*cm, y_pos, f"• Margen de corte (kerf): {margen_mm} mm")
    
    y_pos -= 30
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y_pos, "Resumen por Tablero:")
    y_pos -= 15
    c.setFont("Helvetica-Bold", 9)
    c.drawString(2.5*cm, y_pos, "Tablero")
    c.drawString(5*cm, y_pos, "Piezas")
    c.drawString(8*cm, y_pos, "Área usada")
    c.drawString(12*cm, y_pos, "Desperdicio")
    c.drawString(16*cm, y_pos, "% Uso")
    y_pos -= 5
    c.line(2*cm, y_pos, width - 2*cm, y_pos)
    y_pos -= 12
    c.setFont("Helvetica", 9)

    if info_desperdicio is None:
        info_desperdicio = _info_desperdicio_desde_optimizacion(optimizacion) or {}
    info_tableros = info_desperdicio.get('info_tableros') or []

    for i in range(len(imagenes_base64)):
        if y_pos < 100:
            c.showPage()
            y_pos = height - 50
        c.drawString(2.5*cm, y_pos, f"Tablero {i+1}")
        if i < len(info_tableros):
            row = info_tableros[i]
            n_p = row.get('num_piezas', 0)
            au = round(float(row['area_usada']) * factor_area, 2)
            dp = round(float(row['desperdicio']) * factor_area, 2)
            pct = row.get('porcentaje_uso')
            c.drawString(5*cm, y_pos, str(n_p))
            c.drawString(8*cm, y_pos, f"{au} {simbolo_area}")
            c.drawString(12*cm, y_pos, f"{dp} {simbolo_area}")
            if pct is not None:
                c.drawString(16*cm, y_pos, f"{float(pct):.2f}%")
            else:
                c.drawString(16*cm, y_pos, "-")
        else:
            for col_x in (5*cm, 8*cm, 12*cm, 16*cm):
                c.drawString(col_x, y_pos, "-")
        y_pos -= 12

    # === PÁGINAS SIGUIENTES: Un tablero por página ===
    for i, img_base64 in enumerate(imagenes_base64, start=1):
        if not img_base64 or img_base64.isspace():
            continue

        try:
            c.showPage()
            
            c.setFont("Helvetica-Bold", 16)
            c.drawCentredString(width / 2, height - 40, f"Tablero {i} de {len(imagenes_base64)}")

            image_data = base64.b64decode(img_base64)
            img_temp = os.path.join(settings.MEDIA_ROOT, f"temp_img_{optimizacion.id}_{i}.png")
            with open(img_temp, "wb") as f:
                f.write(image_data)

            with Image.open(img_temp) as im:
                img_width, img_height = im.size
                max_width = 18 * cm
                max_height = 23 * cm
                ratio = min(max_width / img_width, max_height / img_height)
                final_width = img_width * ratio
                final_height = img_height * ratio

            x_pos = (width - final_width) / 2
            y_pos = (height - final_height - 3*cm) / 2
            
            c.drawImage(img_temp, x_pos, y_pos, width=final_width, height=final_height, 
                       preserveAspectRatio=True, mask='auto')

            c.setFont("Helvetica-Oblique", 9)
            c.drawCentredString(width / 2, 1.5*cm, f"Página {i+1} de {len(imagenes_base64)+1}")

            try:
                os.remove(img_temp)
            except Exception:
                pass

        except Exception as e:
            c.setFont("Helvetica-Oblique", 10)
            c.drawString(2*cm, height / 2, f"Error: {str(e)}")

    # Pie de página final
    c.setStrokeColorRGB(0.3, 0.3, 0.3)
    c.line(2*cm, 1*cm, width - 2*cm, 1*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(width / 2, 0.7*cm, "CutLess - Optimización de Recursos en Carpintería | Algoritmo FFD")
    
    c.save()
    
    return os.path.join("pdfs", filename)


def generar_grafico_aprovechamiento(optimizaciones, periodo='todos', alta_resolucion=False):
    """
    Genera un gráfico de línea mostrando la tendencia de aprovechamiento.
    
    Args:
        optimizaciones: QuerySet de optimizaciones ordenadas por fecha
        periodo: 'semanal', 'mensual', 'anual', 'todos'
        alta_resolucion: Si es True, genera una versión en alta resolución para zoom
    
    Returns:
        String base64 de la imagen del gráfico
    """
    if not optimizaciones.exists():
        return None
    
    # Ajustar tamaño y DPI según resolución
    if alta_resolucion:
        fig, ax = plt.subplots(figsize=(16, 9))
        dpi = 200
        fontsize_title = 18
        fontsize_labels = 14
        fontsize_legend = 12
        linewidth = 3
        markersize = 8
    else:
        fig, ax = plt.subplots(figsize=(12, 6))
        dpi = 100
        fontsize_title = 14
        fontsize_labels = 11
        fontsize_legend = 10
        linewidth = 2
        markersize = 6
    
    # Preparar datos
    fechas = [opt.fecha for opt in optimizaciones]
    aprovechamientos = [opt.aprovechamiento_total for opt in optimizaciones]
    
    # Calcular promedio
    promedio = sum(aprovechamientos) / len(aprovechamientos) if aprovechamientos else 0
    
    # Gráfico de línea
    ax.plot(fechas, aprovechamientos, marker='o', linewidth=linewidth, markersize=markersize, 
            color='#4ECDC4', label='Aprovechamiento')
    
    # Línea de promedio
    ax.axhline(y=promedio, color='#FF6B6B', linestyle='--', linewidth=linewidth, 
               label=f'Promedio: {promedio:.2f}%')
    
    # Formatear fechas según período
    if periodo == 'semanal':
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
    elif periodo == 'mensual':
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%Y'))
    elif periodo == 'anual':
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    else:
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m/%Y'))
    
    plt.xticks(rotation=45, ha='right')
    ax.set_xlabel('Fecha', fontsize=fontsize_labels, fontweight='bold')
    ax.set_ylabel('Aprovechamiento (%)', fontsize=fontsize_labels, fontweight='bold')
    ax.set_title('📊 Tendencia de Aprovechamiento', fontsize=fontsize_title, fontweight='bold', pad=20)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.legend(loc='best', fontsize=fontsize_legend)
    ax.set_ylim(0, 100)
    
    plt.tight_layout()
    
    # Convertir a base64
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


def generar_grafico_desperdicio(optimizaciones, periodo='todos', alta_resolucion=False):
    """
    Genera un gráfico de línea mostrando la tendencia de desperdicio.
    Calcula el desperdicio basado en el aprovechamiento.
    
    Args:
        optimizaciones: QuerySet de optimizaciones ordenadas por fecha
        periodo: 'semanal', 'mensual', 'anual', 'todos'
        alta_resolucion: Si es True, genera una versión en alta resolución para zoom
    
    Returns:
        String base64 de la imagen del gráfico
    """
    if not optimizaciones.exists():
        return None
    
    # Ajustar tamaño y DPI según resolución
    if alta_resolucion:
        fig, ax = plt.subplots(figsize=(16, 9))
        dpi = 200
        fontsize_title = 18
        fontsize_labels = 14
        fontsize_legend = 12
        linewidth = 3
        markersize = 8
    else:
        fig, ax = plt.subplots(figsize=(12, 6))
        dpi = 100
        fontsize_title = 14
        fontsize_labels = 11
        fontsize_legend = 10
        linewidth = 2
        markersize = 6
    
    # Preparar datos
    fechas = [opt.fecha for opt in optimizaciones]
    # Calcular desperdicio: 100 - aprovechamiento
    desperdicios = [100 - opt.aprovechamiento_total for opt in optimizaciones]
    
    # Calcular promedio
    promedio = sum(desperdicios) / len(desperdicios) if desperdicios else 0
    
    # Gráfico de línea
    ax.plot(fechas, desperdicios, marker='s', linewidth=linewidth, markersize=markersize, 
            color='#FF6B6B', label='Desperdicio')
    
    # Línea de promedio
    ax.axhline(y=promedio, color='#4ECDC4', linestyle='--', linewidth=linewidth, 
               label=f'Promedio: {promedio:.2f}%')
    
    # Formatear fechas según período
    if periodo == 'semanal':
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
    elif periodo == 'mensual':
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%Y'))
    elif periodo == 'anual':
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    else:
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m/%Y'))
    
    plt.xticks(rotation=45, ha='right')
    ax.set_xlabel('Fecha', fontsize=fontsize_labels, fontweight='bold')
    ax.set_ylabel('Desperdicio (%)', fontsize=fontsize_labels, fontweight='bold')
    ax.set_title('📉 Tendencia de Desperdicio', fontsize=fontsize_title, fontweight='bold', pad=20)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.legend(loc='best', fontsize=fontsize_legend)
    ax.set_ylim(0, 100)
    
    plt.tight_layout()
    
    # Convertir a base64
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


# ===== FUNCIÓN DE EXPORTACIÓN A EXCEL =====

def generar_excel(optimizacion, info_desperdicio, piezas_con_nombre, numero_lista=None):
    """
    Genera un archivo Excel con la información detallada de la optimización.
    
    Args:
        optimizacion: Objeto Optimizacion
        info_desperdicio: Diccionario con información de desperdicio
        piezas_con_nombre: Lista de diccionarios con información de piezas
        numero_lista: Número de lista de la optimización (opcional)
    
    Returns:
        BytesIO: Buffer con el archivo Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # === HOJA 1: RESUMEN GENERAL ===
    ws1 = wb.active
    ws1.title = "Resumen"
    
    # Título
    ws1['A1'] = f"Optimización #{numero_lista if numero_lista else optimizacion.id}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:B1')
    
    # Información general
    row = 3
    # Calcular porcentaje de desperdicio
    area_total_disponible = info_desperdicio.get('area_usada_total', 0) + info_desperdicio.get('desperdicio_total', 0)
    porcentaje_desperdicio = (info_desperdicio.get('desperdicio_total', 0) / area_total_disponible * 100) if area_total_disponible > 0 else 0
    
    datos_generales = [
        ['Fecha:', optimizacion.fecha.strftime('%d/%m/%Y %H:%M')],
        ['Hora:', optimizacion.fecha.strftime('%H:%M') + ' (Chile)'],
        ['Tablero:', f"{round(convertir_desde_cm(optimizacion.ancho_tablero, optimizacion.unidad_medida), 2)} × {round(convertir_desde_cm(optimizacion.alto_tablero, optimizacion.unidad_medida), 2)} {obtener_simbolo_unidad(optimizacion.unidad_medida)}"],
        ['Aprovechamiento:', f"{optimizacion.aprovechamiento_total:.2f}%"],
        ['Área Utilizada:', f"{info_desperdicio.get('area_usada_total', 0)} {obtener_simbolo_area(optimizacion.unidad_medida)}"],
        ['Desperdicio Total:', f"{info_desperdicio.get('desperdicio_total', 0)} {obtener_simbolo_area(optimizacion.unidad_medida)}"],
        ['% Desperdicio:', f"{porcentaje_desperdicio:.2f}%"],
        ['Total Tableros:', len(info_desperdicio.get('info_tableros', []))],
        ['Rotación Automática:', 'Sí' if getattr(optimizacion, 'permitir_rotacion', True) else 'No'],
        ['Margen de Corte:', f"{round(getattr(optimizacion, 'margen_corte', 0.3) * 10, 1)} mm"],
    ]
    
    for label, valor in datos_generales:
        ws1[f'A{row}'] = label
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = valor
        row += 1
    
    # Información de costos (si existe)
    costo_total = optimizacion.get_costo_total()
    if costo_total or optimizacion.precio_tablero or optimizacion.mano_obra:
        row += 1
        ws1[f'A{row}'] = "--- Información de Costos ---"
        ws1[f'A{row}'].font = Font(bold=True, size=12)
        ws1[f'A{row}'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        row += 1
        
        if optimizacion.material:
            ws1[f'A{row}'] = 'Material:'
            ws1[f'A{row}'].font = Font(bold=True)
            ws1[f'B{row}'] = optimizacion.material.nombre
            row += 1
        
        if optimizacion.precio_tablero:
            ws1[f'A{row}'] = 'Precio por Tablero:'
            ws1[f'A{row}'].font = Font(bold=True)
            ws1[f'B{row}'] = f"${optimizacion.precio_tablero:,.0f}"
            row += 1
            
            if optimizacion.num_tableros:
                costo_material = optimizacion.precio_tablero * optimizacion.num_tableros
                ws1[f'A{row}'] = f'Costo de Material ({optimizacion.num_tableros} tableros):'
                ws1[f'A{row}'].font = Font(bold=True)
                ws1[f'B{row}'] = f"${costo_material:,.0f}"
                row += 1
        
        if optimizacion.mano_obra:
            ws1[f'A{row}'] = 'Mano de Obra:'
            ws1[f'A{row}'].font = Font(bold=True)
            ws1[f'B{row}'] = f"${optimizacion.mano_obra:,.0f}"
            row += 1
        
        if costo_total:
            ws1[f'A{row}'] = 'COSTO TOTAL:'
            ws1[f'A{row}'].font = Font(bold=True, size=12)
            ws1[f'A{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            ws1[f'B{row}'] = f"${costo_total:,.0f}"
            ws1[f'B{row}'].font = Font(bold=True, size=12)
            ws1[f'B{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            row += 1
    
    # Ajustar ancho de columnas
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 30
    
    # === HOJA 2: LISTA DE PIEZAS ===
    ws2 = wb.create_sheet("Piezas")
    
    # Encabezados
    headers = ['Nombre', 'Ancho', 'Alto', 'Cantidad', 'Área Unitaria', 'Área Total']
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos de piezas
    unidad_opt = getattr(optimizacion, 'unidad_medida', 'cm') or 'cm'
    simbolo_area = obtener_simbolo_area(unidad_opt)
    factor_lineal = convertir_desde_cm(1, unidad_opt)
    factor_area = factor_lineal ** 2
    
    total_area_piezas = 0
    total_cantidad = 0
    
    for idx, pieza in enumerate(piezas_con_nombre, start=2):
        nombre = pieza.get('nombre', f"Pieza {idx-1}")
        ancho = float(pieza['ancho'])
        alto = float(pieza['alto'])
        cantidad = int(pieza['cantidad'])
        
        # Convertir a cm para cálculos
        ancho_cm = convertir_a_cm(ancho, unidad_opt)
        alto_cm = convertir_a_cm(alto, unidad_opt)
        area_unitaria_cm2 = ancho_cm * alto_cm
        area_total_cm2 = area_unitaria_cm2 * cantidad
        
        # Convertir para mostrar
        area_unitaria_mostrar = round(area_unitaria_cm2 * factor_area, 2)
        area_total_mostrar = round(area_total_cm2 * factor_area, 2)
        
        ws2.cell(row=idx, column=1, value=nombre).border = border
        ws2.cell(row=idx, column=2, value=ancho).border = border
        ws2.cell(row=idx, column=3, value=alto).border = border
        ws2.cell(row=idx, column=4, value=cantidad).border = border
        ws2.cell(row=idx, column=5, value=f"{area_unitaria_mostrar} {simbolo_area}").border = border
        ws2.cell(row=idx, column=6, value=f"{area_total_mostrar} {simbolo_area}").border = border
        
        total_area_piezas += area_total_cm2
        total_cantidad += cantidad
    
    # Fila de totales
    row_total = len(piezas_con_nombre) + 2
    ws2.cell(row=row_total, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=row_total, column=4, value=total_cantidad).font = Font(bold=True)
    ws2.cell(row=row_total, column=6, value=f"{round(total_area_piezas * factor_area, 2)} {simbolo_area}").font = Font(bold=True)
    
    # Ajustar ancho de columnas
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    
    # === HOJA 3: RESUMEN POR TABLERO ===
    ws3 = wb.create_sheet("Resumen por Tablero")
    
    # Estilos adicionales para desperdicio
    good_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # Verde claro
    warning_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Amarillo
    bad_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # Rojo claro
    
    # Encabezados
    headers = ['Tablero', 'Piezas', 'Área Disponible', 'Área Usada', 'Desperdicio', '% Desperdicio', '% Uso']
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos de tableros
    info_tableros = info_desperdicio.get('info_tableros', [])
    for idx, info in enumerate(info_tableros, start=2):
        area_disponible = info.get('area_usada', 0) + info.get('desperdicio', 0)
        porcentaje_desperdicio = (info.get('desperdicio', 0) / area_disponible * 100) if area_disponible > 0 else 0
        porcentaje_uso = info.get('porcentaje_uso', 0)
        
        ws3.cell(row=idx, column=1, value=f"Tablero {info.get('numero', idx-1)}").border = border
        ws3.cell(row=idx, column=1).alignment = center_alignment
        
        ws3.cell(row=idx, column=2, value=info.get('num_piezas', 0)).border = border
        ws3.cell(row=idx, column=2).alignment = center_alignment
        
        ws3.cell(row=idx, column=3, value=f"{area_disponible:.2f} {simbolo_area}").border = border
        ws3.cell(row=idx, column=3).alignment = center_alignment
        
        ws3.cell(row=idx, column=4, value=f"{info.get('area_usada', 0):.2f} {simbolo_area}").border = border
        ws3.cell(row=idx, column=4).alignment = center_alignment
        
        ws3.cell(row=idx, column=5, value=f"{info.get('desperdicio', 0):.2f} {simbolo_area}").border = border
        ws3.cell(row=idx, column=5).alignment = center_alignment
        
        # Columna % Desperdicio con colores
        cell_desperdicio = ws3.cell(row=idx, column=6, value=f"{porcentaje_desperdicio:.2f}%")
        cell_desperdicio.border = border
        cell_desperdicio.alignment = center_alignment
        if porcentaje_desperdicio < 20:
            cell_desperdicio.fill = good_fill
        elif porcentaje_desperdicio < 40:
            cell_desperdicio.fill = warning_fill
        else:
            cell_desperdicio.fill = bad_fill
        
        # Columna % Uso con colores
        cell_uso = ws3.cell(row=idx, column=7, value=f"{porcentaje_uso:.2f}%")
        cell_uso.border = border
        cell_uso.alignment = center_alignment
        if porcentaje_uso >= 80:
            cell_uso.fill = good_fill
        elif porcentaje_uso >= 60:
            cell_uso.fill = warning_fill
        else:
            cell_uso.fill = bad_fill
    
    # Fila de totales
    row_total = len(info_tableros) + 2
    ws3.cell(row=row_total, column=1, value="TOTALES").font = Font(bold=True, size=12)
    ws3.cell(row=row_total, column=1).border = border
    ws3.cell(row=row_total, column=1).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    total_area_disponible = sum(info.get('area_usada', 0) + info.get('desperdicio', 0) for info in info_tableros)
    total_area_usada = sum(info.get('area_usada', 0) for info in info_tableros)
    total_desperdicio = sum(info.get('desperdicio', 0) for info in info_tableros)
    total_porcentaje_desperdicio = (total_desperdicio / total_area_disponible * 100) if total_area_disponible > 0 else 0
    promedio_uso = sum(info.get('porcentaje_uso', 0) for info in info_tableros) / len(info_tableros) if info_tableros else 0
    
    ws3.cell(row=row_total, column=2, value=sum(info.get('num_piezas', 0) for info in info_tableros)).font = Font(bold=True)
    ws3.cell(row=row_total, column=2).border = border
    ws3.cell(row=row_total, column=2).alignment = center_alignment
    ws3.cell(row=row_total, column=2).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    ws3.cell(row=row_total, column=3, value=f"{total_area_disponible:.2f} {simbolo_area}").font = Font(bold=True)
    ws3.cell(row=row_total, column=3).border = border
    ws3.cell(row=row_total, column=3).alignment = center_alignment
    ws3.cell(row=row_total, column=3).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    ws3.cell(row=row_total, column=4, value=f"{total_area_usada:.2f} {simbolo_area}").font = Font(bold=True)
    ws3.cell(row=row_total, column=4).border = border
    ws3.cell(row=row_total, column=4).alignment = center_alignment
    ws3.cell(row=row_total, column=4).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    ws3.cell(row=row_total, column=5, value=f"{total_desperdicio:.2f} {simbolo_area}").font = Font(bold=True)
    ws3.cell(row=row_total, column=5).border = border
    ws3.cell(row=row_total, column=5).alignment = center_alignment
    ws3.cell(row=row_total, column=5).fill = bad_fill
    
    cell_total_desperdicio_pct = ws3.cell(row=row_total, column=6, value=f"{total_porcentaje_desperdicio:.2f}%")
    cell_total_desperdicio_pct.font = Font(bold=True, size=12)
    cell_total_desperdicio_pct.border = border
    cell_total_desperdicio_pct.alignment = center_alignment
    if total_porcentaje_desperdicio < 20:
        cell_total_desperdicio_pct.fill = good_fill
    elif total_porcentaje_desperdicio < 40:
        cell_total_desperdicio_pct.fill = warning_fill
    else:
        cell_total_desperdicio_pct.fill = bad_fill
    
    cell_total_uso = ws3.cell(row=row_total, column=7, value=f"{promedio_uso:.2f}%")
    cell_total_uso.font = Font(bold=True, size=12)
    cell_total_uso.border = border
    cell_total_uso.alignment = center_alignment
    if promedio_uso >= 80:
        cell_total_uso.fill = good_fill
    elif promedio_uso >= 60:
        cell_total_uso.fill = warning_fill
    else:
        cell_total_uso.fill = bad_fill
    
    # Ajustar ancho de columnas
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 18
    ws3.column_dimensions['F'].width = 16
    ws3.column_dimensions['G'].width = 12
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def generar_excel_historial_costos(optimizaciones_con_costo, estadisticas, fecha_desde=None, fecha_hasta=None):
    """
    Genera un archivo Excel con el resumen completo del historial de costos.
    
    Args:
        optimizaciones_con_costo: Lista de diccionarios con optimizaciones y costos
        estadisticas: Diccionario con estadísticas calculadas
        fecha_desde: Fecha desde (opcional)
        fecha_hasta: Fecha hasta (opcional)
    
    Returns:
        BytesIO: Buffer con el archivo Excel
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # === HOJA 1: RESUMEN DE ESTADÍSTICAS ===
    ws1 = wb.active
    ws1.title = "Resumen"
    
    # Título
    ws1['A1'] = "💰 HISTORIAL DE COSTOS - RESUMEN"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:B1')
    
    # Período
    row = 3
    if fecha_desde or fecha_hasta:
        periodo_texto = "Período: "
        if fecha_desde:
            periodo_texto += f"Desde {fecha_desde}"
        if fecha_desde and fecha_hasta:
            periodo_texto += " hasta "
        if fecha_hasta:
            periodo_texto += f"{fecha_hasta}"
        ws1[f'A{row}'] = periodo_texto
        ws1[f'A{row}'].font = Font(bold=True, italic=True)
        ws1.merge_cells(f'A{row}:B{row}')
        row += 2
    else:
        ws1[f'A{row}'] = "Período: Todos los registros"
        ws1[f'A{row}'].font = Font(bold=True, italic=True)
        ws1.merge_cells(f'A{row}:B{row}')
        row += 2
    
    # Estadísticas principales
    stats_data = [
        ['Total Optimizaciones:', estadisticas.get('total_optimizaciones', 0)],
        ['Costo Total:', f"${estadisticas.get('costo_total', 0):,.0f}"],
        ['Costo Material:', f"${estadisticas.get('costo_material_total', 0):,.0f}"],
        ['Mano de Obra:', f"${estadisticas.get('costo_mano_obra_total', 0):,.0f}"],
        ['Total Tableros:', estadisticas.get('num_tableros_total', 0)],
        ['Costo Promedio:', f"${estadisticas.get('costo_promedio', 0):,.0f}"],
        ['Costo por Tablero:', f"${estadisticas.get('costo_por_tablero', 0):,.0f}"],
        ['% Material:', f"{estadisticas.get('porcentaje_material', 0):.1f}%"],
        ['% Mano de Obra:', f"{estadisticas.get('porcentaje_mano_obra', 0):.1f}%"],
        ['Aprovechamiento Promedio:', f"{estadisticas.get('aprovechamiento_promedio', 0):.1f}%"],
    ]
    
    for label, valor in stats_data:
        ws1[f'A{row}'] = label
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = valor
        if 'Costo Total' in label:
            ws1[f'B{row}'].font = Font(bold=True, size=12, color="006100")
        row += 1
    
    # Ajustar ancho de columnas
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20
    
    # === HOJA 2: DETALLE DE OPTIMIZACIONES ===
    ws2 = wb.create_sheet("Detalle de Costos")
    
    # Encabezados
    headers = ['Fecha', 'Hora', 'Dimensiones', 'Material', 'Tableros', 'Aprovechamiento', 
               'Costo Material', 'Mano de Obra', 'Costo Total']
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos de optimizaciones
    for idx, item in enumerate(optimizaciones_con_costo, start=2):
        opt = item['optimizacion']
        
        # Fecha y hora
        fecha_str = opt.fecha.strftime('%d/%m/%Y')
        hora_str = opt.fecha.strftime('%H:%M')
        
        # Dimensiones
        dim_str = f"{opt.ancho_tablero} × {opt.alto_tablero} cm"
        if opt.unidad_medida != 'cm':
            dim_str += f" ({opt.get_unidad_medida_display()})"
        
        # Material
        material_str = opt.material.nombre if opt.material else "-"
        
        # Tableros
        num_tableros = opt.num_tableros or 0
        
        # Aprovechamiento
        aprovechamiento = opt.aprovechamiento_total or 0
        
        # Costos
        costo_material = item['costo_material']
        costo_mano_obra = item['costo_mano_obra']
        costo_total = item['costo_total']
        
        # Escribir datos
        ws2.cell(row=idx, column=1, value=fecha_str).border = border
        ws2.cell(row=idx, column=2, value=hora_str).border = border
        ws2.cell(row=idx, column=3, value=dim_str).border = border
        ws2.cell(row=idx, column=4, value=material_str).border = border
        ws2.cell(row=idx, column=5, value=num_tableros).border = border
        ws2.cell(row=idx, column=6, value=f"{aprovechamiento:.1f}%").border = border
        ws2.cell(row=idx, column=7, value=f"${costo_material:,.0f}").border = border
        ws2.cell(row=idx, column=8, value=f"${costo_mano_obra:,.0f}").border = border
        ws2.cell(row=idx, column=9, value=f"${costo_total:,.0f}").border = border
        
        # Formato de números
        ws2.cell(row=idx, column=5).alignment = center_alignment
        ws2.cell(row=idx, column=6).alignment = center_alignment
        ws2.cell(row=idx, column=7).alignment = Alignment(horizontal='right')
        ws2.cell(row=idx, column=8).alignment = Alignment(horizontal='right')
        ws2.cell(row=idx, column=9).alignment = Alignment(horizontal='right')
        ws2.cell(row=idx, column=9).font = Font(bold=True)
    
    # Fila de totales
    row_total = len(optimizaciones_con_costo) + 2
    ws2.cell(row=row_total, column=1, value="TOTALES:").font = Font(bold=True, size=12)
    ws2.cell(row=row_total, column=1).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    ws2.merge_cells(f'A{row_total}:F{row_total}')
    
    ws2.cell(row=row_total, column=7, value=f"${estadisticas.get('costo_material_total', 0):,.0f}").font = Font(bold=True)
    ws2.cell(row=row_total, column=7).border = border
    ws2.cell(row=row_total, column=7).alignment = Alignment(horizontal='right')
    ws2.cell(row=row_total, column=7).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    ws2.cell(row=row_total, column=8, value=f"${estadisticas.get('costo_mano_obra_total', 0):,.0f}").font = Font(bold=True)
    ws2.cell(row=row_total, column=8).border = border
    ws2.cell(row=row_total, column=8).alignment = Alignment(horizontal='right')
    ws2.cell(row=row_total, column=8).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    ws2.cell(row=row_total, column=9, value=f"${estadisticas.get('costo_total', 0):,.0f}").font = Font(bold=True, size=12, color="006100")
    ws2.cell(row=row_total, column=9).border = border
    ws2.cell(row=row_total, column=9).alignment = Alignment(horizontal='right')
    ws2.cell(row=row_total, column=9).fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    
    # Ajustar ancho de columnas
    column_widths = [12, 8, 18, 20, 10, 15, 15, 15, 15]
    for col, width in enumerate(column_widths, start=1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def generar_pdf_presupuesto(presupuesto):
    """
    Genera un PDF profesional del presupuesto.
    
    Args:
        presupuesto: Objeto Presupuesto
        
    Returns:
        str: Ruta del archivo PDF generado
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from django.utils import timezone
    
    filename = f"presupuesto_{presupuesto.numero.replace('-', '_')}.pdf"
    filepath = os.path.join(settings.MEDIA_ROOT, "presupuestos", filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    c = canvas.Canvas(filepath, pagesize=A4)
    width, height = A4
    
    # === ENCABEZADO ===
    c.setFont("Helvetica-Bold", 24)
    c.setFillColorRGB(0, 0.3, 0.6)
    c.drawCentredString(width / 2, height - 50, "PRESUPUESTO")
    
    c.setFont("Helvetica-Bold", 14)
    c.setFillColorRGB(0, 0, 0)
    c.drawCentredString(width / 2, height - 75, f"Número: {presupuesto.numero}")
    
    # === INFORMACIÓN DEL CLIENTE ===
    y_pos = height - 120
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y_pos, "Cliente:")
    c.setFont("Helvetica", 11)
    y_pos -= 20
    
    if presupuesto.cliente:
        cliente = presupuesto.cliente
        c.drawString(2.5*cm, y_pos, f"• Nombre: {cliente.nombre}")
        y_pos -= 15
        if cliente.rut:
            c.drawString(2.5*cm, y_pos, f"• RUT: {cliente.rut}")
            y_pos -= 15
        if cliente.email:
            c.drawString(2.5*cm, y_pos, f"• Email: {cliente.email}")
            y_pos -= 15
        if cliente.telefono:
            c.drawString(2.5*cm, y_pos, f"• Teléfono: {cliente.telefono}")
            y_pos -= 15
        if cliente.direccion:
            c.drawString(2.5*cm, y_pos, f"• Dirección: {cliente.direccion}")
            y_pos -= 15
    else:
        c.drawString(2.5*cm, y_pos, "• Sin cliente asignado")
        y_pos -= 15
    
    y_pos -= 10
    
    # === INFORMACIÓN DEL PRESUPUESTO ===
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y_pos, "Información del Presupuesto:")
    c.setFont("Helvetica", 11)
    y_pos -= 20
    
    c.drawString(2.5*cm, y_pos, f"• Fecha de creación: {presupuesto.fecha_creacion.strftime('%d/%m/%Y %H:%M')}")
    y_pos -= 15
    c.drawString(2.5*cm, y_pos, f"• Válido hasta: {presupuesto.fecha_validez.strftime('%d/%m/%Y')}")
    y_pos -= 15
    c.drawString(2.5*cm, y_pos, f"• Estado: {presupuesto.get_estado_display()}")
    y_pos -= 15
    
    # === DETALLES DE LAS OPTIMIZACIONES ===
    optimizaciones = presupuesto.optimizaciones.all()
    y_pos -= 20
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y_pos, f"Detalles de las Optimizaciones ({optimizaciones.count()}):")
    c.setFont("Helvetica", 11)
    y_pos -= 20
    
    for idx, optimizacion in enumerate(optimizaciones, start=1):
        unidad = getattr(optimizacion, 'unidad_medida', 'cm') or 'cm'
        ancho_mostrar = convertir_desde_cm(optimizacion.ancho_tablero, unidad)
        alto_mostrar = convertir_desde_cm(optimizacion.alto_tablero, unidad)
        simbolo = obtener_simbolo_unidad(unidad)
        
        c.drawString(2.5*cm, y_pos, f"Optimización #{idx}:")
        y_pos -= 15
        c.drawString(3*cm, y_pos, f"• Dimensiones: {ancho_mostrar:.2f} × {alto_mostrar:.2f} {simbolo}")
        y_pos -= 15
        c.drawString(3*cm, y_pos, f"• Tableros: {optimizacion.num_tableros or 0}")
        y_pos -= 15
        c.drawString(3*cm, y_pos, f"• Aprovechamiento: {optimizacion.aprovechamiento_total:.2f}%")
        y_pos -= 15
        
        if optimizacion.material:
            c.drawString(3*cm, y_pos, f"• Material: {optimizacion.material.nombre}")
            y_pos -= 15
        
        # Agregar lista de piezas con nombres
        if optimizacion.piezas:
            c.setFont("Helvetica-Bold", 10)
            c.drawString(3*cm, y_pos, "• Piezas:")
            y_pos -= 12
            c.setFont("Helvetica", 9)
            
            # Parsear piezas
            piezas_list = []
            for linea in optimizacion.piezas.splitlines():
                if linea.strip():
                    partes = linea.split(',')
                    if len(partes) == 4:  # Formato con nombre: nombre,ancho,alto,cantidad
                        nombre = partes[0].strip()
                        ancho_pieza = float(partes[1].strip())
                        alto_pieza = float(partes[2].strip())
                        cantidad = int(partes[3].strip())
                        # Convertir dimensiones a la unidad de la optimización
                        ancho_mostrar_pieza = round(convertir_desde_cm(ancho_pieza, unidad), 2)
                        alto_mostrar_pieza = round(convertir_desde_cm(alto_pieza, unidad), 2)
                        piezas_list.append({
                            'nombre': nombre,
                            'ancho': ancho_mostrar_pieza,
                            'alto': alto_mostrar_pieza,
                            'cantidad': cantidad
                        })
                    elif len(partes) == 3:  # Formato sin nombre: ancho,alto,cantidad
                        ancho_pieza = float(partes[0].strip())
                        alto_pieza = float(partes[1].strip())
                        cantidad = int(partes[2].strip())
                        ancho_mostrar_pieza = round(convertir_desde_cm(ancho_pieza, unidad), 2)
                        alto_mostrar_pieza = round(convertir_desde_cm(alto_pieza, unidad), 2)
                        piezas_list.append({
                            'nombre': f"Pieza {len(piezas_list) + 1}",
                            'ancho': ancho_mostrar_pieza,
                            'alto': alto_mostrar_pieza,
                            'cantidad': cantidad
                        })
            
            # Mostrar piezas (máximo 5 para no ocupar mucho espacio)
            for pieza in piezas_list[:5]:
                if y_pos < 150:  # Si se acerca al final de la página, salir
                    break
                texto_pieza = f"  - {pieza['nombre']}: {pieza['ancho']:.2f} × {pieza['alto']:.2f} {simbolo} (x{pieza['cantidad']})"
                # Truncar si es muy largo
                if len(texto_pieza) > 70:
                    texto_pieza = texto_pieza[:67] + "..."
                c.drawString(3.5*cm, y_pos, texto_pieza)
                y_pos -= 11
            
            # Si hay más de 5 piezas, indicarlo
            if len(piezas_list) > 5:
                c.drawString(3.5*cm, y_pos, f"  ... y {len(piezas_list) - 5} pieza(s) más")
                y_pos -= 11
            
            c.setFont("Helvetica", 11)  # Volver al tamaño normal
        
        y_pos -= 5  # Espacio entre optimizaciones
    
    # === DESGLOSE DE COSTOS ===
    y_pos -= 20
    c.setFont("Helvetica-Bold", 12)
    c.setFillColorRGB(0, 0.5, 0.8)
    c.drawString(2*cm, y_pos, "Desglose de Costos:")
    c.setFillColorRGB(0, 0, 0)
    c.setFont("Helvetica", 11)
    y_pos -= 20
    
    # Tabla de costos - considerar todas las optimizaciones
    data = [['Concepto', 'Cantidad', 'Precio Unitario', 'Subtotal']]
    
    # Calcular número de lista para cada optimización (basado en orden por fecha descendente)
    todas_optimizaciones_usuario = Optimizacion.objects.filter(usuario=presupuesto.optimizaciones.first().usuario).order_by('-fecha')
    total_optimizaciones_usuario = todas_optimizaciones_usuario.count()
    
    # Agregar fila por cada optimización
    total_tableros = 0
    for optimizacion in optimizaciones:
        num_tableros = optimizacion.num_tableros or 0
        total_tableros += num_tableros
        costo_tableros = presupuesto.precio_tablero * Decimal(str(num_tableros))
        
        # Calcular el número de lista (basado en orden por fecha descendente)
        numero_lista = optimizacion.pk  # Por defecto usar el ID
        for idx, opt in enumerate(todas_optimizaciones_usuario, start=1):
            if opt.id == optimizacion.id:
                # Para orden descendente: numero = total - idx + 1
                numero_lista = total_optimizaciones_usuario - idx + 1
                break
        
        data.append([
            f'Tableros (Optimización #{numero_lista})',
            str(num_tableros),
            f"${presupuesto.precio_tablero:,.0f}",
            f"${costo_tableros:,.0f}"
        ])
    
    # Agregar mano de obra (una sola vez, no por optimización)
    if optimizaciones.count() > 0:
        data.append([
            'Mano de Obra',
            '1',
            f"${presupuesto.mano_obra:,.0f}",
            f"${presupuesto.mano_obra:,.0f}"
        ])
    
    tabla = Table(data, colWidths=[6*cm, 3*cm, 4*cm, 4*cm])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    
    tabla.wrapOn(c, width - 4*cm, height)
    tabla.drawOn(c, 2*cm, y_pos - 80)
    
    # === TOTAL ===
    y_pos_total = y_pos - 120
    c.setFont("Helvetica-Bold", 14)
    c.setFillColorRGB(0.8, 0.2, 0)
    c.drawString(12*cm, y_pos_total, f"TOTAL: ${presupuesto.costo_total:,.0f}")
    c.setFillColorRGB(0, 0, 0)
    
    # === NOTAS ===
    if presupuesto.notas:
        y_pos_notas = y_pos_total - 40
        c.setFont("Helvetica-Bold", 11)
        c.drawString(2*cm, y_pos_notas, "Notas:")
        c.setFont("Helvetica", 10)
        # Dividir notas en líneas
        notas_lines = presupuesto.notas.split('\n')
        y_pos_notas -= 15
        for line in notas_lines[:10]:  # Máximo 10 líneas
            if y_pos_notas < 100:
                break
            c.drawString(2.5*cm, y_pos_notas, line[:80])  # Máximo 80 caracteres por línea
            y_pos_notas -= 12
    
    # === PIE DE PÁGINA ===
    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    c.drawCentredString(width / 2, 30, f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} - CutLess")
    
    c.save()
    return filepath


def generar_excel_resumen_desperdicio(optimizaciones, estadisticas, periodo='todos'):
    """
    Genera un archivo Excel con el resumen de desperdicio desde estadísticas.
    
    Args:
        optimizaciones: QuerySet de optimizaciones filtradas
        estadisticas: Diccionario con estadísticas calculadas
        periodo: Período seleccionado (todos, semanal, mensual, anual)
    
    Returns:
        BytesIO: Buffer con el archivo Excel
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # === HOJA 1: RESUMEN DE DESPERDICIO ===
    ws1 = wb.active
    ws1.title = "Resumen Desperdicio"
    
    # Título
    ws1['A1'] = "📊 RESUMEN DE DESPERDICIO - REPORTE DE EFICIENCIA"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:C1')
    
    # Período
    row = 3
    periodo_texto = {
        'todos': 'Todos los tiempos',
        'semanal': 'Última semana',
        'mensual': 'Último mes',
        'anual': 'Último año'
    }.get(periodo, 'Todos los tiempos')
    ws1[f'A{row}'] = f"Período: {periodo_texto}"
    ws1[f'A{row}'].font = Font(bold=True, italic=True)
    ws1.merge_cells(f'A{row}:C{row}')
    row += 2
    
    # Estadísticas principales
    stats_data = [
        ['Total Optimizaciones:', estadisticas.get('total_optimizaciones', 0)],
        ['Aprovechamiento Promedio:', f"{estadisticas.get('promedio_aprovechamiento', 0):.2f}%"],
        ['Desperdicio Promedio:', f"{estadisticas.get('promedio_desperdicio', 0):.2f}%"],
        ['Mejor Aprovechamiento:', f"{estadisticas.get('max_aprovechamiento', 0):.2f}%"],
        ['Peor Aprovechamiento:', f"{estadisticas.get('min_aprovechamiento', 0):.2f}%"],
    ]
    
    for label, valor in stats_data:
        ws1[f'A{row}'] = label
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = valor
        if 'Desperdicio' in label:
            ws1[f'B{row}'].font = Font(bold=True, size=12, color="DC3545")
        row += 1
    
    # Ajustar ancho de columnas
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 20
    
    # === HOJA 2: DETALLE DE OPTIMIZACIONES ===
    ws2 = wb.create_sheet("Detalle Optimizaciones")
    
    # Encabezados
    headers = ['Fecha', 'Hora', 'Dimensiones', 'Aprovechamiento', 'Desperdicio', 
               'Tableros', 'Área Total', 'Área Usada', 'Área Desperdiciada']
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos de optimizaciones
    for idx, opt in enumerate(optimizaciones.order_by('fecha'), start=2):
        # Fecha y hora
        fecha_str = opt.fecha.strftime('%d/%m/%Y')
        hora_str = opt.fecha.strftime('%H:%M')
        
        # Dimensiones
        unidad_opt = getattr(opt, 'unidad_medida', 'cm') or 'cm'
        ancho_mostrar = round(convertir_desde_cm(opt.ancho_tablero, unidad_opt), 2)
        alto_mostrar = round(convertir_desde_cm(opt.alto_tablero, unidad_opt), 2)
        simbolo = obtener_simbolo_unidad(unidad_opt)
        dim_str = f"{ancho_mostrar} × {alto_mostrar} {simbolo}"
        
        # Aprovechamiento y desperdicio
        aprovechamiento = opt.aprovechamiento_total or 0
        desperdicio = 100 - aprovechamiento
        
        # Tableros
        num_tableros = opt.num_tableros or 1
        
        # Calcular áreas (necesitamos regenerar el gráfico para obtener info_desperdicio)
        # Por ahora usamos valores aproximados basados en aprovechamiento
        area_tablero = opt.ancho_tablero * opt.alto_tablero  # en cm²
        area_total = area_tablero * num_tableros
        area_usada = area_total * (aprovechamiento / 100)
        area_desperdiciada = area_total - area_usada
        
        simbolo_area = obtener_simbolo_area(unidad_opt)
        
        # Escribir datos
        ws2.cell(row=idx, column=1, value=fecha_str).border = border
        ws2.cell(row=idx, column=2, value=hora_str).border = border
        ws2.cell(row=idx, column=3, value=dim_str).border = border
        ws2.cell(row=idx, column=4, value=f"{aprovechamiento:.2f}%").border = border
        ws2.cell(row=idx, column=5, value=f"{desperdicio:.2f}%").border = border
        ws2.cell(row=idx, column=6, value=num_tableros).border = border
        ws2.cell(row=idx, column=7, value=f"{area_total:.2f} {simbolo_area}").border = border
        ws2.cell(row=idx, column=8, value=f"{area_usada:.2f} {simbolo_area}").border = border
        ws2.cell(row=idx, column=9, value=f"{area_desperdiciada:.2f} {simbolo_area}").border = border
        
        # Formato de números
        ws2.cell(row=idx, column=4).alignment = center_alignment
        ws2.cell(row=idx, column=5).alignment = center_alignment
        ws2.cell(row=idx, column=6).alignment = center_alignment
        
        # Colorear según desperdicio
        if desperdicio > 30:
            fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        elif desperdicio > 15:
            fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        else:
            fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        ws2.cell(row=idx, column=5).fill = fill
    
    # Fila de totales
    row_total = len(optimizaciones) + 2
    ws2.cell(row=row_total, column=1, value="TOTALES:").font = Font(bold=True, size=12)
    ws2.cell(row=row_total, column=1).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    ws2.merge_cells(f'A{row_total}:C{row_total}')
    
    # Ajustar ancho de columnas
    column_widths = [12, 8, 18, 15, 15, 10, 15, 15, 18]
    for col, width in enumerate(column_widths, start=1):
        ws2.column_dimensions[get_column_letter(col)].width = width
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def generar_pdf_resumen_desperdicio(optimizaciones, estadisticas, periodo='todos'):
    """
    Genera un PDF con el resumen de desperdicio desde estadísticas.
    
    Args:
        optimizaciones: QuerySet de optimizaciones filtradas
        estadisticas: Diccionario con estadísticas calculadas
        periodo: Período seleccionado (todos, semanal, mensual, anual)
    
    Returns:
        str: Ruta del archivo PDF generado
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from django.utils import timezone
    import os
    from django.conf import settings
    
    filename = f"resumen_desperdicio_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(settings.MEDIA_ROOT, "pdfs", filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    # Retornar path relativo para compatibilidad con FileResponse
    relative_path = os.path.join("pdfs", filename)
    
    c = canvas.Canvas(filepath, pagesize=A4)
    width, height = A4
    
    # === PÁGINA 1: RESUMEN ===
    c.setFont("Helvetica-Bold", 18)
    c.setFillColorRGB(0.86, 0.21, 0.27)  # Rojo para desperdicio
    c.drawCentredString(width / 2, height - 50, "📊 RESUMEN DE DESPERDICIO")
    c.setFont("Helvetica-Bold", 14)
    c.setFillColorRGB(0, 0, 0)
    c.drawCentredString(width / 2, height - 70, "Reporte de Eficiencia - CutLess")
    
    # Período
    periodo_texto = {
        'todos': 'Todos los tiempos',
        'semanal': 'Última semana',
        'mensual': 'Último mes',
        'anual': 'Último año'
    }.get(periodo, 'Todos los tiempos')
    
    c.setFont("Helvetica", 11)
    c.drawString(2*cm, height - 100, f"Período: {periodo_texto}")
    c.drawString(2*cm, height - 115, f"Fecha del reporte: {timezone.now().strftime('%d/%m/%Y %H:%M')}")
    
    # Estadísticas principales
    y_pos = height - 150
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y_pos, "Estadísticas Generales:")
    c.setFont("Helvetica", 11)
    y_pos -= 20
    
    stats = [
        f"Total Optimizaciones: {estadisticas.get('total_optimizaciones', 0)}",
        f"Aprovechamiento Promedio: {estadisticas.get('promedio_aprovechamiento', 0):.2f}%",
        f"Desperdicio Promedio: {estadisticas.get('promedio_desperdicio', 0):.2f}%",
        f"Mejor Aprovechamiento: {estadisticas.get('max_aprovechamiento', 0):.2f}%",
        f"Peor Aprovechamiento: {estadisticas.get('min_aprovechamiento', 0):.2f}%",
    ]
    
    for stat in stats:
        c.drawString(2.5*cm, y_pos, f"• {stat}")
        if 'Desperdicio' in stat:
            c.setFillColorRGB(0.86, 0.21, 0.27)
            c.drawString(2.5*cm, y_pos, f"• {stat}")
            c.setFillColorRGB(0, 0, 0)
        y_pos -= 18
    
    # === PÁGINA 2+: DETALLE ===
    c.showPage()
    
    # Encabezados de tabla
    y_pos = height - 50
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y_pos, "Detalle de Optimizaciones:")
    y_pos -= 25
    
    # Encabezados de columnas
    headers = ['Fecha', 'Dimensiones', 'Aprovech.', 'Desperd.', 'Tableros']
    x_positions = [2*cm, 5*cm, 10*cm, 13*cm, 16*cm]
    
    c.setFont("Helvetica-Bold", 9)
    for header, x_pos in zip(headers, x_positions):
        c.drawString(x_pos, y_pos, header)
    
    y_pos -= 15
    c.setFont("Helvetica", 8)
    
    # Datos
    for opt in optimizaciones.order_by('fecha')[:30]:  # Máximo 30 por página
        if y_pos < 50:
            c.showPage()
            y_pos = height - 50
        
        fecha_str = opt.fecha.strftime('%d/%m/%Y')
        unidad_opt = getattr(opt, 'unidad_medida', 'cm') or 'cm'
        ancho_mostrar = round(convertir_desde_cm(opt.ancho_tablero, unidad_opt), 2)
        alto_mostrar = round(convertir_desde_cm(opt.alto_tablero, unidad_opt), 2)
        simbolo = obtener_simbolo_unidad(unidad_opt)
        dim_str = f"{ancho_mostrar}×{alto_mostrar}{simbolo}"
        
        aprovechamiento = opt.aprovechamiento_total or 0
        desperdicio = 100 - aprovechamiento
        num_tableros = opt.num_tableros or 1
        
        c.drawString(x_positions[0], y_pos, fecha_str)
        c.drawString(x_positions[1], y_pos, dim_str)
        c.drawString(x_positions[2], y_pos, f"{aprovechamiento:.1f}%")
        
        # Colorear desperdicio
        if desperdicio > 30:
            c.setFillColorRGB(0.86, 0.21, 0.27)
        elif desperdicio > 15:
            c.setFillColorRGB(1.0, 0.76, 0.03)
        else:
            c.setFillColorRGB(0.13, 0.55, 0.13)
        c.drawString(x_positions[3], y_pos, f"{desperdicio:.1f}%")
        c.setFillColorRGB(0, 0, 0)
        
        c.drawString(x_positions[4], y_pos, str(num_tableros))
        
        y_pos -= 15
    
    # Pie de página
    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0.5, 0.5, 0.5)
    c.drawCentredString(width / 2, 30, f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} - CutLess")
    
    c.save()
    return relative_path