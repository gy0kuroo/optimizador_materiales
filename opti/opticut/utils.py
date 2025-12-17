# Imports estándar de Python
import io
import os
import base64

# Imports de terceros
import matplotlib
matplotlib.use('Agg')
import matplotlib.dates as mdates
import matplotlib.patches as patches
import matplotlib.pyplot as plt
from PIL import Image
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas

# Imports de Django
from django.conf import settings


# ===== FUNCIONES DE CONVERSIÓN DE UNIDADES =====
# Todas las conversiones son a centímetros (unidad base)

def convertir_a_cm(valor, unidad):
    """
    Convierte un valor de cualquier unidad a centímetros.
    
    Args:
        valor: Valor numérico a convertir
        unidad: Unidad de origen ('cm', 'm', 'in', 'ft')
    
    Returns:
        Valor convertido a centímetros
    """
    conversiones = {
        'cm': 1.0,           # 1 cm = 1 cm
        'm': 100.0,          # 1 m = 100 cm
        'in': 2.54,          # 1 pulgada = 2.54 cm
        'ft': 30.48,         # 1 pie = 30.48 cm
    }
    factor = conversiones.get(unidad, 1.0)
    return round(valor * factor, 2)


def convertir_desde_cm(valor_cm, unidad_destino):
    """
    Convierte un valor en centímetros a otra unidad.
    
    Args:
        valor_cm: Valor en centímetros
        unidad_destino: Unidad de destino ('cm', 'm', 'in', 'ft')
    
    Returns:
        Valor convertido a la unidad destino
    """
    conversiones = {
        'cm': 1.0,           # 1 cm = 1 cm
        'm': 0.01,           # 1 cm = 0.01 m
        'in': 1/2.54,        # 1 cm = 1/2.54 pulgadas
        'ft': 1/30.48,       # 1 cm = 1/30.48 pies
    }
    factor = conversiones.get(unidad_destino, 1.0)
    return round(valor_cm * factor, 2)


def obtener_simbolo_unidad(unidad):
    """
    Retorna el símbolo de la unidad para mostrar.
    """
    simbolos = {
        'cm': 'cm',
        'm': 'm',
        'in': 'in',
        'ft': 'ft',
    }
    return simbolos.get(unidad, 'cm')


def obtener_simbolo_area(unidad):
    """
    Retorna el símbolo de unidad al cuadrado para áreas.
    """
    simbolos = {
        'cm': 'cm²',
        'm': 'm²',
        'in': 'in²',
        'ft': 'ft²',
    }
    return simbolos.get(unidad, 'cm²')


def generar_grafico(piezas, ancho_tablero, alto_tablero, unidad='cm', permitir_rotacion=True, margen_corte=0.3, nombres_piezas=None):
    """
    Algoritmo First Fit Decreasing (FFD) mejorado para optimización de cortes.
    Incluye rotación automática de piezas y margen de corte (kerf).
    
    Args:
        piezas: Lista de tuplas (ancho, alto, cantidad) en cm
        ancho_tablero: Ancho del tablero en cm
        alto_tablero: Alto del tablero en cm
        unidad: Unidad de medida para mostrar
        permitir_rotacion: Si True, intenta rotar piezas 90° para mejor aprovechamiento
        margen_corte: Margen de corte (kerf) en cm. Se suma al espacio necesario entre piezas
        nombres_piezas: Lista opcional de nombres (no usado actualmente, para compatibilidad)
    """
    AREA_TABLERO = ancho_tablero * alto_tablero
    tableros = []
    area_usada_total = 0

    # Paso 1: Expandir piezas y ordenarlas (FFD)
    piezas_expandidas = []
    for w, h, c in piezas:
        for _ in range(c):
            area = w * h
            piezas_expandidas.append({
                'ancho': w,
                'alto': h,
                'area': area,
                'original': (w, h),
                'rotada': False
            })
    
    piezas_expandidas.sort(key=lambda x: x['area'], reverse=True)
    
    # Función auxiliar para calcular desperdicio estimado
    def calcular_desperdicio_estimado(tablero, ancho, alto):
        """Calcula el desperdicio estimado de un tablero"""
        area_usada = 0
        for pos in tablero['posiciones']:
            if len(pos) >= 5:
                _, _, w, h, _ = pos[:5]  # Obtener w y h
                area_usada += w * h
        return AREA_TABLERO - area_usada
    
    # Función auxiliar para intentar colocar una pieza en un tablero
    def intentar_colocar_pieza(pieza, tablero, w, h, rotada=False):
        """
        Intenta colocar una pieza de dimensiones w x h en el tablero.
        Prioriza llenar niveles existentes con la MISMA altura para maximizar aprovechamiento.
        """
        posiciones = tablero['posiciones']
        niveles = tablero['niveles']
        w_orig = pieza.get('ancho', w)
        h_orig = pieza.get('alto', h)
        
        mejor_nivel_idx = None
        mejor_prioridad = float('inf')
        
        # Buscar el mejor nivel existente
        for idx, nivel in enumerate(niveles):
            x = nivel['x_actual']
            altura_nivel = nivel['altura']
            espacio_horizontal = ancho_tablero - x
            
            # La pieza debe caber
            if h > altura_nivel or w > espacio_horizontal:
                continue
            
            # Prioridad: preferir niveles donde la altura coincide exactamente
            # Esto maximiza el aprovechamiento al evitar espacios verticales perdidos
            diferencia_altura = abs(altura_nivel - h)
            
            # Dar máxima prioridad a altura exacta (0), luego por espacio sobrante
            if diferencia_altura < mejor_prioridad:
                mejor_prioridad = diferencia_altura
                mejor_nivel_idx = idx
        
        # Si encontramos un nivel adecuado, colocar ahí
        if mejor_nivel_idx is not None:
            nivel = niveles[mejor_nivel_idx]
            x = nivel['x_actual']
            y = nivel['y_inicio']
            posiciones.append((x, y, w, h, rotada, w_orig, h_orig))
            nivel['x_actual'] += w + margen_corte
            return True
        
        # Si no cabe en niveles existentes, crear nuevo nivel
        if niveles:
            ultimo_nivel = niveles[-1]
            y_nuevo_nivel = ultimo_nivel['y_inicio'] + ultimo_nivel['altura'] + margen_corte
            
            if y_nuevo_nivel + h <= alto_tablero and w <= ancho_tablero:
                niveles.append({
                    'y_inicio': y_nuevo_nivel,
                    'x_actual': w + margen_corte,
                    'altura': h
                })
                posiciones.append((0, y_nuevo_nivel, w, h, rotada, w_orig, h_orig))
                return True
        
        return False
    
    # Paso 2: Algoritmo de empaquetado por niveles con rotación (MEJORADO: Best Fit)
    # Diccionario para recordar la mejor orientación por tamaño de pieza
    orientacion_optima = {}  # {(w, h): rotada}
    
    for pieza in piezas_expandidas:
        w_original = pieza['ancho']
        h_original = pieza['alto']
        colocada = False
        
        # Si la pieza no es cuadrada y se permite rotación, probar ambas orientaciones
        if permitir_rotacion and w_original != h_original:
            tam_pieza = (w_original, h_original)
            
            # Si ya determinamos la orientación óptima para este tamaño, usarla
            if tam_pieza in orientacion_optima:
                usar_rotacion_previa = orientacion_optima[tam_pieza]
                if usar_rotacion_previa:
                    w_usar, h_usar = h_original, w_original
                else:
                    w_usar, h_usar = w_original, h_original
                
                # Intentar colocar con la orientación predeterminada
                for tablero_idx, tablero in enumerate(tableros):
                    tablero_temp = {
                        'posiciones': list(tablero['posiciones']),
                        'niveles': [dict(n) for n in tablero['niveles']]
                    }
                    if intentar_colocar_pieza(pieza, tablero_temp, w_usar, h_usar, rotada=usar_rotacion_previa):
                        tableros[tablero_idx] = tablero_temp
                        area_usada_total += w_original * h_original
                        pieza['rotada'] = usar_rotacion_previa
                        colocada = True
                        break
            
            # Si no se colocó (primera pieza de este tamaño o no cabe), buscar mejor opción
            if not colocada:
                mejor_resultado = None
                mejor_tablero_idx = None
                mejor_score = float('inf')
                mejor_rotada = False
                
                for tablero_idx, tablero in enumerate(tableros):
                    for nivel in tablero['niveles']:
                        altura_nivel = nivel['altura']
                        espacio_h = ancho_tablero - nivel['x_actual']
                        
                        # Probar orientación original
                        if h_original <= altura_nivel and w_original <= espacio_h:
                            tablero_temp = {
                                'posiciones': list(tablero['posiciones']),
                                'niveles': [dict(n) for n in tablero['niveles']]
                            }
                            if intentar_colocar_pieza(pieza, tablero_temp, w_original, h_original, rotada=False):
                                diff_altura = abs(altura_nivel - h_original)
                                desperdicio = calcular_desperdicio_estimado(tablero_temp, ancho_tablero, alto_tablero)
                                score = diff_altura * 10000 + desperdicio
                                if score < mejor_score:
                                    mejor_score = score
                                    mejor_resultado = tablero_temp
                                    mejor_tablero_idx = tablero_idx
                                    mejor_rotada = False
                        
                        # Probar orientación rotada
                        if w_original <= altura_nivel and h_original <= espacio_h:
                            tablero_temp = {
                                'posiciones': list(tablero['posiciones']),
                                'niveles': [dict(n) for n in tablero['niveles']]
                            }
                            if intentar_colocar_pieza(pieza, tablero_temp, h_original, w_original, rotada=True):
                                diff_altura = abs(altura_nivel - w_original)
                                desperdicio = calcular_desperdicio_estimado(tablero_temp, ancho_tablero, alto_tablero)
                                score = diff_altura * 10000 + desperdicio
                                if score < mejor_score:
                                    mejor_score = score
                                    mejor_resultado = tablero_temp
                                    mejor_tablero_idx = tablero_idx
                                    mejor_rotada = True
                    
                    # Probar nuevo nivel
                    if tablero['niveles']:
                        ultimo = tablero['niveles'][-1]
                        y_nuevo = ultimo['y_inicio'] + ultimo['altura'] + margen_corte
                        
                        if y_nuevo + h_original <= alto_tablero and w_original <= ancho_tablero:
                            tablero_temp = {
                                'posiciones': list(tablero['posiciones']),
                                'niveles': [dict(n) for n in tablero['niveles']]
                            }
                            if intentar_colocar_pieza(pieza, tablero_temp, w_original, h_original, rotada=False):
                                desperdicio = calcular_desperdicio_estimado(tablero_temp, ancho_tablero, alto_tablero)
                                if desperdicio < mejor_score:
                                    mejor_score = desperdicio
                                    mejor_resultado = tablero_temp
                                    mejor_tablero_idx = tablero_idx
                                    mejor_rotada = False
                        
                        if y_nuevo + w_original <= alto_tablero and h_original <= ancho_tablero:
                            tablero_temp = {
                                'posiciones': list(tablero['posiciones']),
                                'niveles': [dict(n) for n in tablero['niveles']]
                            }
                            if intentar_colocar_pieza(pieza, tablero_temp, h_original, w_original, rotada=True):
                                desperdicio = calcular_desperdicio_estimado(tablero_temp, ancho_tablero, alto_tablero)
                                if desperdicio < mejor_score:
                                    mejor_score = desperdicio
                                    mejor_resultado = tablero_temp
                                    mejor_tablero_idx = tablero_idx
                                    mejor_rotada = True
                
                if mejor_resultado is not None:
                    tableros[mejor_tablero_idx] = mejor_resultado
                    area_usada_total += w_original * h_original
                    pieza['rotada'] = mejor_rotada
                    orientacion_optima[tam_pieza] = mejor_rotada  # Recordar orientación
                    colocada = True
        else:
            # Sin rotación o pieza cuadrada: MEJORA - probar TODOS los tableros (Best Fit)
            mejor_tablero_idx = None
            mejor_desperdicio = float('inf')
            mejor_tablero = None
            
            for tablero_idx, tablero in enumerate(tableros):
                tablero_temp = {
                    'posiciones': list(tablero['posiciones']),
                    'niveles': [dict(n) for n in tablero['niveles']]
                }
                if intentar_colocar_pieza(pieza, tablero_temp, w_original, h_original, rotada=False):
                    desperdicio = calcular_desperdicio_estimado(tablero_temp, ancho_tablero, alto_tablero)
                    if desperdicio < mejor_desperdicio:
                        mejor_desperdicio = desperdicio
                        mejor_tablero = tablero_temp
                        mejor_tablero_idx = tablero_idx
            
            if mejor_tablero is not None:
                tableros[mejor_tablero_idx] = mejor_tablero
                area_usada_total += w_original * h_original
                colocada = True
        
        # Si no se pudo colocar en ningún tablero existente, crear uno nuevo
        if not colocada:
            if permitir_rotacion and w_original != h_original:
                # Definir las dos orientaciones posibles
                # Orientación A: pieza tal como viene (w_original x h_original)
                # Orientación B: pieza rotada (h_original x w_original)
                w_a, h_a = w_original, h_original
                w_b, h_b = h_original, w_original
                
                # Verificar si cada orientación cabe en el tablero
                cabe_a = w_a <= ancho_tablero and h_a <= alto_tablero
                cabe_b = w_b <= ancho_tablero and h_b <= alto_tablero
                
                if not cabe_a and not cabe_b:
                    continue
                
                # Calcular cuántas piezas cabrían en cada orientación
                if cabe_a:
                    piezas_x_a = int(ancho_tablero / (w_a + margen_corte)) or (1 if w_a <= ancho_tablero else 0)
                    piezas_y_a = int(alto_tablero / (h_a + margen_corte)) or (1 if h_a <= alto_tablero else 0)
                    # Verificar que realmente caben
                    if piezas_x_a * (w_a + margen_corte) - margen_corte > ancho_tablero:
                        piezas_x_a = max(1, piezas_x_a - 1)
                    if piezas_y_a * (h_a + margen_corte) - margen_corte > alto_tablero:
                        piezas_y_a = max(1, piezas_y_a - 1)
                    total_a = piezas_x_a * piezas_y_a
                else:
                    total_a = 0
                
                if cabe_b:
                    piezas_x_b = int(ancho_tablero / (w_b + margen_corte)) or (1 if w_b <= ancho_tablero else 0)
                    piezas_y_b = int(alto_tablero / (h_b + margen_corte)) or (1 if h_b <= alto_tablero else 0)
                    # Verificar que realmente caben
                    if piezas_x_b * (w_b + margen_corte) - margen_corte > ancho_tablero:
                        piezas_x_b = max(1, piezas_x_b - 1)
                    if piezas_y_b * (h_b + margen_corte) - margen_corte > alto_tablero:
                        piezas_y_b = max(1, piezas_y_b - 1)
                    total_b = piezas_x_b * piezas_y_b
                else:
                    total_b = 0
                
                # Elegir la orientación que permite más piezas
                # Si son iguales, preferir no rotar
                usar_rotada = total_b > total_a
                
                if usar_rotada:
                    nuevo_tablero = {
                        'posiciones': [(0, 0, w_b, h_b, True, w_original, h_original)],
                        'niveles': [{
                            'y_inicio': 0,
                            'x_actual': w_b + margen_corte,
                            'altura': h_b
                        }]
                    }
                    pieza['rotada'] = True
                    orientacion_optima[(w_original, h_original)] = True  # Recordar
                else:
                    nuevo_tablero = {
                        'posiciones': [(0, 0, w_a, h_a, False, w_original, h_original)],
                        'niveles': [{
                            'y_inicio': 0,
                            'x_actual': w_a + margen_corte,
                            'altura': h_a
                        }]
                    }
                    orientacion_optima[(w_original, h_original)] = False  # Recordar
            else:
                # Sin rotación o pieza cuadrada
                if w_original + margen_corte <= ancho_tablero and h_original <= alto_tablero:
                    # Cabe con margen
                    nuevo_tablero = {
                        'posiciones': [(0, 0, w_original, h_original, False, w_original, h_original)],
                        'niveles': [{
                            'y_inicio': 0,
                            'x_actual': w_original + margen_corte,
                            'altura': h_original
                        }]
                    }
                elif w_original <= ancho_tablero and h_original <= alto_tablero:
                    # Cabe sin margen (última pieza)
                    nuevo_tablero = {
                        'posiciones': [(0, 0, w_original, h_original, False, w_original, h_original)],
                        'niveles': [{
                            'y_inicio': 0,
                            'x_actual': w_original,
                            'altura': h_original
                        }]
                    }
                else:
                    # No cabe - esto no debería pasar si las validaciones son correctas
                    continue
            tableros.append(nuevo_tablero)
            area_usada_total += w_original * h_original
    
    # Paso 3: Calcular aprovechamiento y desperdicio
    num_tableros = len(tableros)
    if num_tableros == 0:
        aprovechamiento_total = 0
        desperdicio_total = 0
        info_tableros = []
    else:
        area_total_disponible = num_tableros * AREA_TABLERO
        desperdicio_total = area_total_disponible - area_usada_total
        aprovechamiento_total = round((area_usada_total / area_total_disponible) * 100, 2)
        
        # Calcular desperdicio por tablero
        info_tableros = []
        for idx, tablero in enumerate(tableros, start=1):
            # Las posiciones ahora incluyen: (x, y, w, h, rotada, w_orig, h_orig) o (x, y, w, h, rotada)
            # Calcular área usando las dimensiones colocadas (w, h)
            area_usada_tablero = 0
            for pos in tablero['posiciones']:
                if len(pos) >= 5:
                    _, _, w, h, _ = pos[:5]  # Obtener w y h (dimensiones colocadas)
                    area_usada_tablero += w * h
            desperdicio_tablero = AREA_TABLERO - area_usada_tablero
            porcentaje_uso = round((area_usada_tablero / AREA_TABLERO) * 100, 2)
            
            info_tableros.append({
                'numero': idx,
                'area_usada': area_usada_tablero,
                'desperdicio': desperdicio_tablero,
                'porcentaje_uso': porcentaje_uso,
                'num_piezas': len(tablero['posiciones'])
            })
    
    # Paso 4: Generar imágenes
    imagenes_base64 = []
    
    for i, tablero in enumerate(tableros, start=1):
        posiciones = tablero['posiciones']
        info = info_tableros[i-1]
        
        fig, ax = plt.subplots(figsize=(10, 10))
        ax.set_xlim(0, ancho_tablero)
        ax.set_ylim(0, alto_tablero)
        ax.invert_yaxis()
        ax.set_aspect('equal')
        
        # Convertir desperdicio para mostrar
        simbolo = obtener_simbolo_unidad(unidad)
        simbolo_area = obtener_simbolo_area(unidad)
        factor_lineal = convertir_desde_cm(1, unidad)
        factor_area = factor_lineal ** 2
        desperdicio_mostrar = round(info['desperdicio'] * factor_area, 2)
        
        # Título con información de desperdicio
        ax.set_title(f"Tablero {i} de {num_tableros} - FFD\n"
                    f"Uso: {info['porcentaje_uso']}% | Desperdicio: {desperdicio_mostrar} {simbolo_area}",
                    fontsize=13, fontweight='bold', pad=20)
        ax.set_xlabel(f"Ancho ({simbolo})", fontsize=11)
        ax.set_ylabel(f"Alto ({simbolo})", fontsize=11)
        
        ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5)
        ax.set_axisbelow(True)
        
        # Borde del tablero
        borde = patches.Rectangle((0, 0), ancho_tablero, alto_tablero,
                                  linewidth=3, edgecolor='black', 
                                  facecolor='#f0f0f0', alpha=0.3)
        ax.add_patch(borde)
        
        # Dibujar piezas (ahora incluyen información de rotación)
        colores = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E2']
        
        for idx, pos_data in enumerate(posiciones):
            # Manejar diferentes formatos de posiciones
            if len(pos_data) >= 7:
                # Formato nuevo: (x, y, w, h, rotada, w_original, h_original)
                x, y, w, h, rotada, w_orig, h_orig = pos_data
            elif len(pos_data) == 5:
                # Formato intermedio: (x, y, w, h, rotada)
                x, y, w, h, rotada = pos_data
                w_orig, h_orig = w, h
            else:
                # Formato antiguo: (x, y, w, h)
                x, y, w, h = pos_data
                rotada = False
                w_orig, h_orig = w, h
            
            color = colores[idx % len(colores)]
            
            rect = patches.Rectangle((x, y), w, h, linewidth=2,
                                     edgecolor='darkblue', facecolor=color, alpha=0.7)
            ax.add_patch(rect)
            
            # Mostrar dimensiones
            w_orig_mostrar = round(convertir_desde_cm(w_orig, unidad), 1)
            h_orig_mostrar = round(convertir_desde_cm(h_orig, unidad), 1)
            
            # Mostrar dimensiones con indicador de rotación
            if rotada:
                texto_dimensiones = f'{w_orig_mostrar}×{h_orig_mostrar}\n(ROTADA 90°)'
            else:
                texto_dimensiones = f'{w_orig_mostrar}×{h_orig_mostrar}'
            
            ax.text(x + w/2, y + h/2, texto_dimensiones,
                   ha='center', va='center', fontsize=9 if rotada else 10, 
                   fontweight='bold', color='white',
                   bbox=dict(boxstyle='round,pad=0.3', facecolor='darkgreen' if rotada else 'black', alpha=0.8))
        
        # Información detallada (convertir áreas)
        area_usada_mostrar = round(info['area_usada'] * factor_area, 2)
        info_text = (f"Piezas: {info['num_piezas']}\n"
                    f"Área usada: {area_usada_mostrar} {simbolo_area}\n"
                    f"Desperdicio: {desperdicio_mostrar} {simbolo_area}")
        ax.text(ancho_tablero * 0.02, alto_tablero * 0.98, info_text,
               fontsize=9, verticalalignment='top',
               bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.9))
        
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=120, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        buf.seek(0)
        imagenes_base64.append(base64.b64encode(buf.read()).decode("utf-8"))
    
    # Retornar también información de desperdicio
    return imagenes_base64, aprovechamiento_total, {
        'area_usada_total': area_usada_total,
        'desperdicio_total': desperdicio_total,
        'info_tableros': info_tableros,
        'num_tableros': num_tableros,
        'area_total_disponible': num_tableros * AREA_TABLERO if num_tableros > 0 else 0
    }

def generar_pdf(optimizacion, imagenes_base64, numero_lista=None):
    """
    Genera UN SOLO PDF con todos los tableros, cada uno en su propia página.
    
    Args:
        optimizacion: Objeto Optimizacion
        imagenes_base64: Lista de imágenes en base64
        numero_lista: Número de la lista en el historial (opcional). Si se proporciona,
                     se usará en el nombre del archivo en lugar del ID.
    """
    if isinstance(imagenes_base64, str):
        imagenes_base64 = [imagenes_base64] if imagenes_base64 else []
    
    if not imagenes_base64:
        return None
    
    # Usar número de lista si está disponible, sino usar el ID
    if numero_lista is not None:
        filename = f"optimizacion_{optimizacion.usuario.username}_{numero_lista}.pdf"
    else:
        filename = f"optimizacion_{optimizacion.usuario.username}_{optimizacion.id}.pdf"
    filepath = os.path.join(settings.MEDIA_ROOT, "pdfs", filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    c = canvas.Canvas(filepath, pagesize=A4)
    width, height = A4

    # === PÁGINA 1: Información general ===
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(width / 2, height - 50, "Reporte de Optimización - OptiCut")
    c.setFont("Helvetica-Oblique", 10)
    c.drawCentredString(width / 2, height - 70, "Algoritmo: First Fit Decreasing (FFD)")

    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, height - 110, "Información General:")
    c.setFont("Helvetica", 11)
    c.drawString(2.5*cm, height - 130, f"• Usuario: {optimizacion.usuario.username}")
    c.drawString(2.5*cm, height - 145, f"• Fecha: {optimizacion.fecha.strftime('%d/%m/%Y %H:%M')}")
    # Obtener unidad y convertir dimensiones para mostrar
    unidad = getattr(optimizacion, 'unidad_medida', 'cm') or 'cm'
    ancho_mostrar = convertir_desde_cm(optimizacion.ancho_tablero, unidad)
    alto_mostrar = convertir_desde_cm(optimizacion.alto_tablero, unidad)
    simbolo = obtener_simbolo_unidad(unidad)
    
    c.drawString(2.5*cm, height - 160, f"• Dimensiones del tablero: {ancho_mostrar} × {alto_mostrar} {simbolo}")
    c.drawString(2.5*cm, height - 175, f"• Tableros generados: {len(imagenes_base64)}")
    
    c.setFont("Helvetica-Bold", 12)
    c.setFillColorRGB(0, 0.5, 0)
    c.drawString(2.5*cm, height - 195, f"• Aprovechamiento: {optimizacion.aprovechamiento_total:.2f}%")
    c.setFillColorRGB(0, 0, 0)

    # Encabezado de lista de piezas
    # Tabla de piezas detallada
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, height - 225, "Lista de Piezas:")
    c.setFont("Helvetica-Bold", 10)
    
    # Encabezados de tabla
    y_pos = height - 245
    c.drawString(2.5*cm, y_pos, "Cant.")
    c.drawString(4*cm, y_pos, "Nombre")
    c.drawString(9*cm, y_pos, "Dimensiones")
    c.drawString(14*cm, y_pos, "Área unit.")
    c.drawString(17.5*cm, y_pos, "Área total")
    
    # Línea separadora
    y_pos -= 5
    c.line(2*cm, y_pos, width - 2*cm, y_pos)
    y_pos -= 15
    
    c.setFont("Helvetica", 9)
    total_area_piezas = 0
    total_cantidad_piezas = 0
    
    for linea in optimizacion.piezas.splitlines():
        if y_pos < 120:  # Más espacio para tabla
            c.showPage()
            y_pos = height - 50
            # Reimprimir encabezados si hay nueva página
            c.setFont("Helvetica-Bold", 10)
            c.drawString(2.5*cm, y_pos, "Cant.")
            c.drawString(4*cm, y_pos, "Nombre")
            c.drawString(9*cm, y_pos, "Dimensiones")
            c.drawString(14*cm, y_pos, "Área unit.")
            c.drawString(17.5*cm, y_pos, "Área total")
            y_pos -= 5
            c.line(2*cm, y_pos, width - 2*cm, y_pos)
            y_pos -= 15
            c.setFont("Helvetica", 9)
        
        try:
            partes = linea.split(',')
            if len(partes) == 4:
                nombre, ancho_str, alto_str, cantidad_str = partes
                nombre = nombre.strip()
                ancho = float(ancho_str.strip())
                alto = float(alto_str.strip())
                cantidad = int(cantidad_str.strip())
            else:
                ancho_str, alto_str, cantidad_str = partes
                nombre = "Pieza"
                ancho = float(ancho_str.strip())
                alto = float(alto_str.strip())
                cantidad = int(cantidad_str.strip())
            
            # Convertir a cm si es necesario
            ancho_cm = convertir_a_cm(ancho, optimizacion.unidad_medida)
            alto_cm = convertir_a_cm(alto, optimizacion.unidad_medida)
            
            # Calcular áreas
            area_unit_cm2 = ancho_cm * alto_cm
            area_total_cm2 = area_unit_cm2 * cantidad
            
            # Convertir para mostrar
            simbolo_area = obtener_simbolo_area(optimizacion.unidad_medida)
            factor_lineal = convertir_desde_cm(1, optimizacion.unidad_medida)
            factor_area = factor_lineal ** 2
            
            area_unit_mostrar = round(area_unit_cm2 * factor_area, 2)
            area_total_mostrar = round(area_total_cm2 * factor_area, 2)
            
            ancho_mostrar = round(convertir_desde_cm(ancho_cm, optimizacion.unidad_medida), 1)
            alto_mostrar = round(convertir_desde_cm(alto_cm, optimizacion.unidad_medida), 1)
            
            # Dibujar fila
            c.drawString(2.5*cm, y_pos, str(cantidad))
            c.drawString(4*cm, y_pos, nombre[:20])  # Limitar longitud
            c.drawString(9*cm, y_pos, f"{ancho_mostrar} × {alto_mostrar} {obtener_simbolo_unidad(optimizacion.unidad_medida)}")
            c.drawString(14*cm, y_pos, f"{area_unit_mostrar} {simbolo_area}")
            c.drawString(17.5*cm, y_pos, f"{area_total_mostrar} {simbolo_area}")
            
            total_area_piezas += area_total_cm2
            total_cantidad_piezas += cantidad
            
        except Exception as e:
            c.drawString(2.5*cm, y_pos, f"Error: {linea}")
        
        y_pos -= 15
    
    # Resumen de piezas
    y_pos -= 10
    c.line(2*cm, y_pos, width - 2*cm, y_pos)
    y_pos -= 15
    c.setFont("Helvetica-Bold", 10)
    simbolo_area = obtener_simbolo_area(optimizacion.unidad_medida)
    factor_lineal = convertir_desde_cm(1, optimizacion.unidad_medida)
    factor_area = factor_lineal ** 2
    total_area_mostrar = round(total_area_piezas * factor_area, 2)
    c.drawString(17.5*cm, y_pos, f"Total: {total_area_mostrar} {simbolo_area}")
    c.drawString(2.5*cm, y_pos, f"Total piezas: {total_cantidad_piezas}")
    
    # Información de configuración
    y_pos -= 30
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y_pos, "Configuración de optimización:")
    c.setFont("Helvetica", 9)
    y_pos -= 15
    rotacion_texto = "Sí" if getattr(optimizacion, 'permitir_rotacion', True) else "No"
    c.drawString(2.5*cm, y_pos, f"• Rotación automática: {rotacion_texto}")
    y_pos -= 12
    # El margen de corte se guarda en cm, pero siempre se muestra en mm
    margen_cm = getattr(optimizacion, 'margen_corte', 0.3)
    margen_mm = round(margen_cm * 10, 1)  # Convertir de cm a mm
    c.drawString(2.5*cm, y_pos, f"• Margen de corte (kerf): {margen_mm} mm")
    
    # Tabla de desperdicio por tablero (si hay información disponible)
    # Esto se calculará desde las imágenes generadas, pero por ahora usamos datos básicos
    y_pos -= 30
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, y_pos, "Resumen por Tablero:")
    y_pos -= 15
    c.setFont("Helvetica-Bold", 9)
    c.drawString(2.5*cm, y_pos, "Tablero")
    c.drawString(5*cm, y_pos, "Piezas")
    c.drawString(8*cm, y_pos, "Área usada")
    c.drawString(12*cm, y_pos, "Desperdicio")
    c.drawString(16*cm, y_pos, "% Uso")
    y_pos -= 5
    c.line(2*cm, y_pos, width - 2*cm, y_pos)
    y_pos -= 12
    c.setFont("Helvetica", 9)
    
    # Calcular info de tableros desde las piezas (aproximado)
    # En una implementación completa, esto vendría de info_desperdicio
    for i in range(len(imagenes_base64)):
        if y_pos < 100:
            c.showPage()
            y_pos = height - 50
        # Información básica - en producción esto vendría de info_desperdicio
        c.drawString(2.5*cm, y_pos, f"Tablero {i+1}")
        c.drawString(5*cm, y_pos, "-")  # Se calcularía desde las posiciones
        c.drawString(8*cm, y_pos, "-")  # Se calcularía desde las posiciones
        c.drawString(12*cm, y_pos, "-")  # Se calcularía desde las posiciones
        c.drawString(16*cm, y_pos, "-")  # Se calcularía desde las posiciones
        y_pos -= 12

    # === PÁGINAS SIGUIENTES: Un tablero por página ===
    for i, img_base64 in enumerate(imagenes_base64, start=1):
        if not img_base64 or img_base64.isspace():
            continue

        try:
            c.showPage()
            
            c.setFont("Helvetica-Bold", 16)
            c.drawCentredString(width / 2, height - 40, f"Tablero {i} de {len(imagenes_base64)}")

            image_data = base64.b64decode(img_base64)
            img_temp = os.path.join(settings.MEDIA_ROOT, f"temp_img_{optimizacion.id}_{i}.png")
            with open(img_temp, "wb") as f:
                f.write(image_data)

            with Image.open(img_temp) as im:
                img_width, img_height = im.size
                max_width = 18 * cm
                max_height = 23 * cm
                ratio = min(max_width / img_width, max_height / img_height)
                final_width = img_width * ratio
                final_height = img_height * ratio

            x_pos = (width - final_width) / 2
            y_pos = (height - final_height - 3*cm) / 2
            
            c.drawImage(img_temp, x_pos, y_pos, width=final_width, height=final_height, 
                       preserveAspectRatio=True, mask='auto')

            c.setFont("Helvetica-Oblique", 9)
            c.drawCentredString(width / 2, 1.5*cm, f"Página {i+1} de {len(imagenes_base64)+1}")

            try:
                os.remove(img_temp)
            except Exception:
                pass

        except Exception as e:
            c.setFont("Helvetica-Oblique", 10)
            c.drawString(2*cm, height / 2, f"Error: {str(e)}")

    # Pie de página final
    c.setStrokeColorRGB(0.3, 0.3, 0.3)
    c.line(2*cm, 1*cm, width - 2*cm, 1*cm)
    c.setFont("Helvetica", 8)
    c.drawCentredString(width / 2, 0.7*cm, "OptiCut - Optimización de Recursos en Carpintería | Algoritmo FFD")
    
    c.save()
    
    return os.path.join("pdfs", filename)


def generar_grafico_aprovechamiento(optimizaciones, periodo='todos', alta_resolucion=False):
    """
    Genera un gráfico de línea mostrando la tendencia de aprovechamiento.
    
    Args:
        optimizaciones: QuerySet de optimizaciones ordenadas por fecha
        periodo: 'semanal', 'mensual', 'anual', 'todos'
        alta_resolucion: Si es True, genera una versión en alta resolución para zoom
    
    Returns:
        String base64 de la imagen del gráfico
    """
    if not optimizaciones.exists():
        return None
    
    # Ajustar tamaño y DPI según resolución
    if alta_resolucion:
        fig, ax = plt.subplots(figsize=(16, 9))
        dpi = 200
        fontsize_title = 18
        fontsize_labels = 14
        fontsize_legend = 12
        linewidth = 3
        markersize = 8
    else:
        fig, ax = plt.subplots(figsize=(12, 6))
        dpi = 100
        fontsize_title = 14
        fontsize_labels = 11
        fontsize_legend = 10
        linewidth = 2
        markersize = 6
    
    # Preparar datos
    fechas = [opt.fecha for opt in optimizaciones]
    aprovechamientos = [opt.aprovechamiento_total for opt in optimizaciones]
    
    # Calcular promedio
    promedio = sum(aprovechamientos) / len(aprovechamientos) if aprovechamientos else 0
    
    # Gráfico de línea
    ax.plot(fechas, aprovechamientos, marker='o', linewidth=linewidth, markersize=markersize, 
            color='#4ECDC4', label='Aprovechamiento')
    
    # Línea de promedio
    ax.axhline(y=promedio, color='#FF6B6B', linestyle='--', linewidth=linewidth, 
               label=f'Promedio: {promedio:.2f}%')
    
    # Formatear fechas según período
    if periodo == 'semanal':
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
    elif periodo == 'mensual':
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%Y'))
    elif periodo == 'anual':
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    else:
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m/%Y'))
    
    plt.xticks(rotation=45, ha='right')
    ax.set_xlabel('Fecha', fontsize=fontsize_labels, fontweight='bold')
    ax.set_ylabel('Aprovechamiento (%)', fontsize=fontsize_labels, fontweight='bold')
    ax.set_title('📊 Tendencia de Aprovechamiento', fontsize=fontsize_title, fontweight='bold', pad=20)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.legend(loc='best', fontsize=fontsize_legend)
    ax.set_ylim(0, 100)
    
    plt.tight_layout()
    
    # Convertir a base64
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


def generar_grafico_desperdicio(optimizaciones, periodo='todos', alta_resolucion=False):
    """
    Genera un gráfico de línea mostrando la tendencia de desperdicio.
    Calcula el desperdicio basado en el aprovechamiento.
    
    Args:
        optimizaciones: QuerySet de optimizaciones ordenadas por fecha
        periodo: 'semanal', 'mensual', 'anual', 'todos'
        alta_resolucion: Si es True, genera una versión en alta resolución para zoom
    
    Returns:
        String base64 de la imagen del gráfico
    """
    if not optimizaciones.exists():
        return None
    
    # Ajustar tamaño y DPI según resolución
    if alta_resolucion:
        fig, ax = plt.subplots(figsize=(16, 9))
        dpi = 200
        fontsize_title = 18
        fontsize_labels = 14
        fontsize_legend = 12
        linewidth = 3
        markersize = 8
    else:
        fig, ax = plt.subplots(figsize=(12, 6))
        dpi = 100
        fontsize_title = 14
        fontsize_labels = 11
        fontsize_legend = 10
        linewidth = 2
        markersize = 6
    
    # Preparar datos
    fechas = [opt.fecha for opt in optimizaciones]
    # Calcular desperdicio: 100 - aprovechamiento
    desperdicios = [100 - opt.aprovechamiento_total for opt in optimizaciones]
    
    # Calcular promedio
    promedio = sum(desperdicios) / len(desperdicios) if desperdicios else 0
    
    # Gráfico de línea
    ax.plot(fechas, desperdicios, marker='s', linewidth=linewidth, markersize=markersize, 
            color='#FF6B6B', label='Desperdicio')
    
    # Línea de promedio
    ax.axhline(y=promedio, color='#4ECDC4', linestyle='--', linewidth=linewidth, 
               label=f'Promedio: {promedio:.2f}%')
    
    # Formatear fechas según período
    if periodo == 'semanal':
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
    elif periodo == 'mensual':
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%m/%Y'))
    elif periodo == 'anual':
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y'))
    else:
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m/%Y'))
    
    plt.xticks(rotation=45, ha='right')
    ax.set_xlabel('Fecha', fontsize=fontsize_labels, fontweight='bold')
    ax.set_ylabel('Desperdicio (%)', fontsize=fontsize_labels, fontweight='bold')
    ax.set_title('📉 Tendencia de Desperdicio', fontsize=fontsize_title, fontweight='bold', pad=20)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.legend(loc='best', fontsize=fontsize_legend)
    ax.set_ylim(0, 100)
    
    plt.tight_layout()
    
    # Convertir a base64
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode('utf-8')


# ===== FUNCIÓN DE EXPORTACIÓN A EXCEL =====

def generar_excel(optimizacion, info_desperdicio, piezas_con_nombre, numero_lista=None):
    """
    Genera un archivo Excel con la información detallada de la optimización.
    
    Args:
        optimizacion: Objeto Optimizacion
        info_desperdicio: Diccionario con información de desperdicio
        piezas_con_nombre: Lista de diccionarios con información de piezas
        numero_lista: Número de lista de la optimización (opcional)
    
    Returns:
        BytesIO: Buffer con el archivo Excel
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # === HOJA 1: RESUMEN GENERAL ===
    ws1 = wb.active
    ws1.title = "Resumen"
    
    # Título
    ws1['A1'] = f"Optimización #{numero_lista if numero_lista else optimizacion.id}"
    ws1['A1'].font = title_font
    ws1.merge_cells('A1:B1')
    
    # Información general
    row = 3
    datos_generales = [
        ['Fecha:', optimizacion.fecha.strftime('%d/%m/%Y %H:%M')],
        ['Hora:', optimizacion.fecha.strftime('%H:%M') + ' (Chile)'],
        ['Tablero:', f"{round(convertir_desde_cm(optimizacion.ancho_tablero, optimizacion.unidad_medida), 2)} × {round(convertir_desde_cm(optimizacion.alto_tablero, optimizacion.unidad_medida), 2)} {obtener_simbolo_unidad(optimizacion.unidad_medida)}"],
        ['Aprovechamiento:', f"{optimizacion.aprovechamiento_total:.2f}%"],
        ['Área Utilizada:', f"{info_desperdicio.get('area_usada_total', 0)} {obtener_simbolo_area(optimizacion.unidad_medida)}"],
        ['Desperdicio Total:', f"{info_desperdicio.get('desperdicio_total', 0)} {obtener_simbolo_area(optimizacion.unidad_medida)}"],
        ['Total Tableros:', len(info_desperdicio.get('info_tableros', []))],
        ['Rotación Automática:', 'Sí' if getattr(optimizacion, 'permitir_rotacion', True) else 'No'],
        ['Margen de Corte:', f"{round(getattr(optimizacion, 'margen_corte', 0.3) * 10, 1)} mm"],
    ]
    
    for label, valor in datos_generales:
        ws1[f'A{row}'] = label
        ws1[f'A{row}'].font = Font(bold=True)
        ws1[f'B{row}'] = valor
        row += 1
    
    # Ajustar ancho de columnas
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 30
    
    # === HOJA 2: LISTA DE PIEZAS ===
    ws2 = wb.create_sheet("Piezas")
    
    # Encabezados
    headers = ['Nombre', 'Ancho', 'Alto', 'Cantidad', 'Área Unitaria', 'Área Total']
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos de piezas
    unidad_opt = getattr(optimizacion, 'unidad_medida', 'cm') or 'cm'
    simbolo_area = obtener_simbolo_area(unidad_opt)
    factor_lineal = convertir_desde_cm(1, unidad_opt)
    factor_area = factor_lineal ** 2
    
    total_area_piezas = 0
    total_cantidad = 0
    
    for idx, pieza in enumerate(piezas_con_nombre, start=2):
        nombre = pieza.get('nombre', f"Pieza {idx-1}")
        ancho = float(pieza['ancho'])
        alto = float(pieza['alto'])
        cantidad = int(pieza['cantidad'])
        
        # Convertir a cm para cálculos
        ancho_cm = convertir_a_cm(ancho, unidad_opt)
        alto_cm = convertir_a_cm(alto, unidad_opt)
        area_unitaria_cm2 = ancho_cm * alto_cm
        area_total_cm2 = area_unitaria_cm2 * cantidad
        
        # Convertir para mostrar
        area_unitaria_mostrar = round(area_unitaria_cm2 * factor_area, 2)
        area_total_mostrar = round(area_total_cm2 * factor_area, 2)
        
        ws2.cell(row=idx, column=1, value=nombre).border = border
        ws2.cell(row=idx, column=2, value=ancho).border = border
        ws2.cell(row=idx, column=3, value=alto).border = border
        ws2.cell(row=idx, column=4, value=cantidad).border = border
        ws2.cell(row=idx, column=5, value=f"{area_unitaria_mostrar} {simbolo_area}").border = border
        ws2.cell(row=idx, column=6, value=f"{area_total_mostrar} {simbolo_area}").border = border
        
        total_area_piezas += area_total_cm2
        total_cantidad += cantidad
    
    # Fila de totales
    row_total = len(piezas_con_nombre) + 2
    ws2.cell(row=row_total, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=row_total, column=4, value=total_cantidad).font = Font(bold=True)
    ws2.cell(row=row_total, column=6, value=f"{round(total_area_piezas * factor_area, 2)} {simbolo_area}").font = Font(bold=True)
    
    # Ajustar ancho de columnas
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 15
    
    # === HOJA 3: RESUMEN POR TABLERO ===
    ws3 = wb.create_sheet("Resumen por Tablero")
    
    # Encabezados
    headers = ['Tablero', 'Piezas', 'Área Usada', 'Desperdicio', '% Uso']
    for col, header in enumerate(headers, start=1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Datos de tableros
    for idx, info in enumerate(info_desperdicio.get('info_tableros', []), start=2):
        ws3.cell(row=idx, column=1, value=f"Tablero {info.get('numero', idx-1)}").border = border
        ws3.cell(row=idx, column=2, value=info.get('num_piezas', 0)).border = border
        ws3.cell(row=idx, column=3, value=f"{info.get('area_usada', 0)} {simbolo_area}").border = border
        ws3.cell(row=idx, column=4, value=f"{info.get('desperdicio', 0)} {simbolo_area}").border = border
        ws3.cell(row=idx, column=5, value=f"{info.get('porcentaje_uso', 0):.2f}%").border = border
    
    # Ajustar ancho de columnas
    for col in range(1, 6):
        ws3.column_dimensions[get_column_letter(col)].width = 18
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer